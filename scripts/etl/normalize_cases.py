"""Normalize raw case issue records.

This step reads the Cases workbook and emits one raw case record per related
equipment ID. It preserves the original System Elements text, Issue Image,
and Corrective Images fields for downstream matching and data quality checks.
"""

from __future__ import annotations

import csv
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

try:
    from .file_discovery import get_input_files
    from .json_utils import file_metadata, write_json as write_json_payload
except ImportError:
    from file_discovery import get_input_files
    from json_utils import file_metadata, write_json as write_json_payload


PROJECT_ROOT = Path(__file__).resolve().parents[2]
OUTPUT_JSON_PATH = PROJECT_ROOT / "frontend" / "public" / "data" / "cases_raw.json"
OUTPUT_CSV_PATH = PROJECT_ROOT / "frontend" / "public" / "data" / "cases_raw.csv"
SOURCE_SHEET_NAME = "EXPORT"

SOURCE_COLUMNS = [
    "Item #",
    "Category",
    "Status",
    "Priority",
    "Summary",
    "System Elements",
    "Issue Image",
    "Reported On",
    "Due",
    "Assigned to",
    "Customer",
    "Contract",
    "Created",
    "Last Updated",
    "Billing Type",
]

OPTIONAL_SOURCE_COLUMNS = [
    "Corrective Images",
]

OUTPUT_FIELDS = [
    "case_id",
    "category",
    "status",
    "priority",
    "summary",
    "system_element_raw",
    "equipment_id",
    "match_status",
    "issue_image",
    "corrective_images",
    "reported_on",
    "due_date",
    "assigned_to",
    "customer",
    "contract",
    "created_at",
    "last_updated_at",
    "billing_type",
]

DATE_TEXT_PATTERNS = [
    re.compile(
        r"\b\d{1,2}/\d{1,2}/\d{2,4}"
        r"(?:\s+\d{1,2}:\d{2}(?::\d{2})?\s*(?:AM|PM)?)?\b",
        re.IGNORECASE,
    ),
    re.compile(
        r"\b\d{4}-\d{1,2}-\d{1,2}"
        r"(?:[ T]\d{1,2}:\d{2}(?::\d{2}(?:\.\d+)?)?)?\b",
        re.IGNORECASE,
    ),
]

DATE_FORMATS = [
    "%m/%d/%Y %I:%M:%S %p",
    "%m/%d/%Y %I:%M %p",
    "%m/%d/%Y %H:%M:%S",
    "%m/%d/%Y %H:%M",
    "%m/%d/%Y",
    "%m/%d/%y %I:%M:%S %p",
    "%m/%d/%y %I:%M %p",
    "%m/%d/%y %H:%M:%S",
    "%m/%d/%y %H:%M",
    "%m/%d/%y",
    "%Y-%m-%dT%H:%M:%S.%f",
    "%Y-%m-%dT%H:%M:%S",
    "%Y-%m-%d %H:%M:%S.%f",
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%d %H:%M",
    "%Y-%m-%d",
]


def is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")


def clean_text(value: Any) -> str | None:
    if is_blank(value):
        return None
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def excel_serial_to_iso(value: int | float) -> str | None:
    if isinstance(value, bool):
        return None
    try:
        parsed = from_excel(value)
    except (TypeError, ValueError, OverflowError):
        return None
    return parsed.isoformat()


def parse_datetime_string(value: str) -> str | None:
    value = value.strip()
    if not value:
        return None

    try:
        return datetime.fromisoformat(value).isoformat()
    except ValueError:
        pass

    for date_format in DATE_FORMATS:
        try:
            return datetime.strptime(value, date_format).isoformat()
        except ValueError:
            continue

    return None


def extract_datetime_string(value: str) -> str | None:
    direct_parse = parse_datetime_string(value)
    if direct_parse is not None:
        return direct_parse

    for pattern in DATE_TEXT_PATTERNS:
        match = pattern.search(value)
        if match is None:
            continue
        parsed = parse_datetime_string(match.group(0))
        if parsed is not None:
            return parsed

    return None


def to_iso_datetime(value: Any) -> str | None:
    if is_blank(value):
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return excel_serial_to_iso(value)
    if isinstance(value, str):
        return extract_datetime_string(value)
    return None


def make_header_map(header_row: tuple[Any, ...]) -> dict[str, int]:
    header_map: dict[str, int] = {}
    for index, header in enumerate(header_row):
        cleaned_header = clean_text(header)
        if cleaned_header is not None:
            header_map[cleaned_header] = index
    return header_map


def validate_source_columns(header_map: dict[str, int]) -> None:
    missing_columns = [
        source_column
        for source_column in SOURCE_COLUMNS
        if source_column not in header_map
    ]
    if missing_columns:
        missing_display = ", ".join(missing_columns)
        raise ValueError(f"Missing expected Cases columns: {missing_display}")


def get_cell(row: tuple[Any, ...], header_map: dict[str, int], source_column: str) -> Any:
    index = header_map[source_column]
    if index >= len(row):
        return None
    return row[index]


def get_optional_cell(
    row: tuple[Any, ...],
    header_map: dict[str, int],
    source_column: str,
) -> Any:
    if source_column not in header_map:
        return None
    return get_cell(row, header_map, source_column)


def row_is_empty(row: tuple[Any, ...], header_map: dict[str, int]) -> bool:
    return all(
        is_blank(get_optional_cell(row, header_map, column))
        for column in [*SOURCE_COLUMNS, *OPTIONAL_SOURCE_COLUMNS]
    )


def split_system_elements(value: Any) -> tuple[str | None, list[str | None]]:
    system_element_raw = clean_text(value)
    if system_element_raw is None:
        return None, [None]

    id_matches = re.findall(r"IAD06-[^;,\n\r]+", system_element_raw, flags=re.IGNORECASE)
    equipment_ids = [match.strip() for match in id_matches if match.strip()]
    if len(equipment_ids) > 1:
        return system_element_raw, equipment_ids

    parts = [
        part.strip()
        for part in re.split(r"[;,\n\r]+", system_element_raw)
        if part.strip()
    ]
    if len(parts) > 1:
        return system_element_raw, parts

    return system_element_raw, [system_element_raw]


def make_case_record(
    row: tuple[Any, ...],
    header_map: dict[str, int],
    system_element_raw: str | None,
    equipment_id: str | None,
) -> dict[str, Any]:
    return {
        "case_id": clean_text(get_cell(row, header_map, "Item #")),
        "category": clean_text(get_cell(row, header_map, "Category")),
        "status": clean_text(get_cell(row, header_map, "Status")),
        "priority": clean_text(get_cell(row, header_map, "Priority")),
        "summary": clean_text(get_cell(row, header_map, "Summary")),
        "system_element_raw": system_element_raw,
        "equipment_id": equipment_id,
        "match_status": "pending_match",
        "issue_image": clean_text(get_cell(row, header_map, "Issue Image")),
        "corrective_images": clean_text(
            get_optional_cell(row, header_map, "Corrective Images")
        ),
        "reported_on": to_iso_datetime(get_cell(row, header_map, "Reported On")),
        "due_date": to_iso_datetime(get_cell(row, header_map, "Due")),
        "assigned_to": clean_text(get_cell(row, header_map, "Assigned to")),
        "customer": clean_text(get_cell(row, header_map, "Customer")),
        "contract": clean_text(get_cell(row, header_map, "Contract")),
        "created_at": to_iso_datetime(get_cell(row, header_map, "Created")),
        "last_updated_at": to_iso_datetime(get_cell(row, header_map, "Last Updated")),
        "billing_type": clean_text(get_cell(row, header_map, "Billing Type")),
    }


def normalize_cases(cases_path: str) -> list[dict]:
    workbook = load_workbook(cases_path, read_only=True, data_only=True)
    try:
        if SOURCE_SHEET_NAME not in workbook.sheetnames:
            raise ValueError(f"Missing expected sheet {SOURCE_SHEET_NAME!r} in {cases_path}")

        sheet = workbook[SOURCE_SHEET_NAME]
        rows = sheet.iter_rows(values_only=True)
        try:
            header_row = next(rows)
        except StopIteration:
            return []

        header_map = make_header_map(header_row)
        validate_source_columns(header_map)

        records: list[dict] = []
        for row in rows:
            if row_is_empty(row, header_map):
                continue

            system_element_raw, equipment_ids = split_system_elements(
                get_cell(row, header_map, "System Elements")
            )
            for equipment_id in equipment_ids:
                records.append(
                    make_case_record(row, header_map, system_element_raw, equipment_id)
                )
    finally:
        workbook.close()

    return records


def write_json(records: list[dict], output_path: Path, source_path: str) -> None:
    write_json_payload(
        output_path,
        {
            "source_file": file_metadata(source_path),
            "records": records,
        },
    )


def write_csv(records: list[dict], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8", newline="") as output_file:
        writer = csv.DictWriter(output_file, fieldnames=OUTPUT_FIELDS)
        writer.writeheader()
        writer.writerows(records)


def main() -> None:
    cases_path = get_input_files()["cases"]
    records = normalize_cases(cases_path)
    write_json(records, OUTPUT_JSON_PATH, cases_path)
    write_csv(records, OUTPUT_CSV_PATH)
    print(f"Wrote {len(records)} case records to {OUTPUT_JSON_PATH}")
    print(f"Wrote {len(records)} case records to {OUTPUT_CSV_PATH}")


if __name__ == "__main__":
    main()

"""Normalize SystemElements equipment records.

This step reads the SystemElements workbook only and emits equipment records
using the normalized schema fields expected by the PDM-centric dashboard.
"""

from __future__ import annotations

import csv
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

try:
    from .file_discovery import get_input_files
    from .json_utils import file_metadata, write_json as write_json_payload
except ImportError:
    from file_discovery import get_input_files
    from json_utils import file_metadata, write_json as write_json_payload


PROJECT_ROOT = Path(__file__).resolve().parents[2]
OUTPUT_JSON_PATH = PROJECT_ROOT / "frontend" / "public" / "data" / "equipment.json"
OUTPUT_CSV_PATH = PROJECT_ROOT / "frontend" / "public" / "data" / "equipment.csv"
SOURCE_SHEET_NAME = "EXPORT"

SOURCE_COLUMNS = [
    "Unique ID",
    "Type",
    "Status",
    "Parent",
    "System",
    "Updated",
    "Updated By",
    "Open Issues",
    "NETA Test Report",
    "NETA Complete: Completed",
    "Equipment Manufacturer",
    "Equipment Model",
    "Equipment Serial Number",
]

OUTPUT_FIELDS = [
    "equipment_id",
    "equipment_type",
    "status",
    "parent",
    "system",
    "open_issues_count_from_system_elements",
    "neta_complete",
    "neta_completed_at",
    "neta_test_report",
    "manufacturer",
    "model",
    "serial_number",
    "updated_at",
    "updated_by",
]

NULL_NETA_MARKERS = {
    "",
    "n/a",
    "na",
    "n.a.",
    "none",
    "null",
    "not applicable",
    "not required",
}
INCOMPLETE_NETA_MARKERS = {
    "false",
    "incomplete",
    "no",
    "not complete",
    "not completed",
    "open",
    "pending",
}
COMPLETE_NETA_MARKERS = {
    "checked",
    "complete",
    "completed",
    "done",
    "pass",
    "passed",
    "true",
    "x",
    "yes",
    "\u2713",
    "\u2714",
}

DATE_TEXT_PATTERNS = [
    re.compile(
        r"\b\d{1,2}/\d{1,2}/\d{2,4}"
        r"(?:\s+\d{1,2}:\d{2}(?::\d{2})?\s*(?:AM|PM)?)?\b",
        re.IGNORECASE,
    ),
    re.compile(
        r"\b\d{4}-\d{1,2}-\d{1,2}"
        r"(?:[ T]\d{1,2}:\d{2}(?::\d{2})?)?\b",
        re.IGNORECASE,
    ),
]

DATE_FORMATS = [
    "%m/%d/%Y %I:%M:%S %p",
    "%m/%d/%Y %I:%M %p",
    "%m/%d/%Y %H:%M:%S",
    "%m/%d/%Y %H:%M",
    "%m/%d/%Y",
    "%m/%d/%y %I:%M:%S %p",
    "%m/%d/%y %I:%M %p",
    "%m/%d/%y %H:%M:%S",
    "%m/%d/%y %H:%M",
    "%m/%d/%y",
    "%Y-%m-%dT%H:%M:%S",
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%d %H:%M",
    "%Y-%m-%d",
]


def is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")


def clean_text(value: Any) -> str | None:
    if is_blank(value):
        return None
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_key(value: str) -> str:
    return " ".join(value.strip().casefold().split())


def contains_marker(value: str, markers: set[str]) -> bool:
    for marker in markers:
        marker_key = normalize_key(marker)
        if marker_key in {"\u2713", "\u2714"} and marker_key in value:
            return True
        if re.search(rf"(?<!\w){re.escape(marker_key)}(?!\w)", value):
            return True
    return False


def excel_serial_to_iso(value: int | float) -> str | None:
    if isinstance(value, bool):
        return None
    try:
        parsed = from_excel(value)
    except (TypeError, ValueError, OverflowError):
        return None
    return parsed.isoformat()


def parse_datetime_string(value: str) -> str | None:
    value = value.strip()
    if not value:
        return None

    try:
        return datetime.fromisoformat(value).isoformat()
    except ValueError:
        pass

    for date_format in DATE_FORMATS:
        try:
            return datetime.strptime(value, date_format).isoformat()
        except ValueError:
            continue

    return None


def extract_datetime_string(value: str) -> str | None:
    direct_parse = parse_datetime_string(value)
    if direct_parse is not None:
        return direct_parse

    for pattern in DATE_TEXT_PATTERNS:
        match = pattern.search(value)
        if match is None:
            continue
        parsed = parse_datetime_string(match.group(0))
        if parsed is not None:
            return parsed

    return None


def to_iso_datetime(value: Any) -> str | None:
    if is_blank(value):
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return excel_serial_to_iso(value)
    if isinstance(value, str):
        return extract_datetime_string(value)
    return None


def parse_open_issues(value: Any) -> int:
    if is_blank(value):
        return 0
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)

    text = str(value).strip().replace(",", "")
    if text == "":
        return 0
    return int(float(text))


def parse_neta_complete(value: Any) -> tuple[bool | None, str | None]:
    if is_blank(value):
        return None, None

    if isinstance(value, bool):
        return value, None

    completed_at = to_iso_datetime(value)
    if completed_at is not None:
        return True, completed_at

    text = str(value).strip()
    normalized = normalize_key(text)

    if normalized in NULL_NETA_MARKERS:
        return None, None
    if normalized in INCOMPLETE_NETA_MARKERS:
        return False, None
    if normalized in COMPLETE_NETA_MARKERS:
        return True, None
    if contains_marker(normalized, INCOMPLETE_NETA_MARKERS):
        return False, None
    if contains_marker(normalized, COMPLETE_NETA_MARKERS):
        return True, None

    return None, None


def make_header_map(header_row: tuple[Any, ...]) -> dict[str, int]:
    header_map: dict[str, int] = {}
    for index, header in enumerate(header_row):
        cleaned_header = clean_text(header)
        if cleaned_header is not None:
            header_map[cleaned_header] = index
    return header_map


def get_cell(row: tuple[Any, ...], header_map: dict[str, int], source_column: str) -> Any:
    index = header_map[source_column]
    if index >= len(row):
        return None
    return row[index]


def row_is_empty(row: tuple[Any, ...], header_map: dict[str, int]) -> bool:
    return all(is_blank(get_cell(row, header_map, column)) for column in SOURCE_COLUMNS)


def normalize_row(row: tuple[Any, ...], header_map: dict[str, int]) -> dict[str, Any]:
    neta_complete, neta_completed_at = parse_neta_complete(
        get_cell(row, header_map, "NETA Complete: Completed")
    )

    return {
        "equipment_id": clean_text(get_cell(row, header_map, "Unique ID")),
        "equipment_type": clean_text(get_cell(row, header_map, "Type")),
        "status": clean_text(get_cell(row, header_map, "Status")),
        "parent": clean_text(get_cell(row, header_map, "Parent")),
        "system": clean_text(get_cell(row, header_map, "System")),
        "open_issues_count_from_system_elements": parse_open_issues(
            get_cell(row, header_map, "Open Issues")
        ),
        "neta_complete": neta_complete,
        "neta_completed_at": neta_completed_at,
        "neta_test_report": clean_text(get_cell(row, header_map, "NETA Test Report")),
        "manufacturer": clean_text(
            get_cell(row, header_map, "Equipment Manufacturer")
        ),
        "model": clean_text(get_cell(row, header_map, "Equipment Model")),
        "serial_number": clean_text(
            get_cell(row, header_map, "Equipment Serial Number")
        ),
        "updated_at": to_iso_datetime(get_cell(row, header_map, "Updated")),
        "updated_by": clean_text(get_cell(row, header_map, "Updated By")),
    }


def validate_source_columns(header_map: dict[str, int]) -> None:
    missing_columns = [
        source_column
        for source_column in SOURCE_COLUMNS
        if source_column not in header_map
    ]
    if missing_columns:
        missing_display = ", ".join(missing_columns)
        raise ValueError(f"Missing expected SystemElements columns: {missing_display}")


def normalize_equipment(system_elements_path: str) -> list[dict]:
    workbook = load_workbook(system_elements_path, read_only=True, data_only=True)
    try:
        if SOURCE_SHEET_NAME not in workbook.sheetnames:
            raise ValueError(
                f"Missing expected sheet {SOURCE_SHEET_NAME!r} in {system_elements_path}"
            )

        sheet = workbook[SOURCE_SHEET_NAME]
        rows = sheet.iter_rows(values_only=True)
        try:
            header_row = next(rows)
        except StopIteration:
            return []

        header_map = make_header_map(header_row)
        validate_source_columns(header_map)

        records = [
            normalize_row(row, header_map)
            for row in rows
            if not row_is_empty(row, header_map)
        ]
    finally:
        workbook.close()

    return records


def write_json(records: list[dict], output_path: Path, source_path: str) -> None:
    write_json_payload(
        output_path,
        {
            "source_file": file_metadata(source_path),
            "records": records,
        },
    )


def write_csv(records: list[dict], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8", newline="") as output_file:
        writer = csv.DictWriter(output_file, fieldnames=OUTPUT_FIELDS)
        writer.writeheader()
        writer.writerows(records)


def main() -> None:
    system_elements_path = get_input_files()["system_elements"]
    records = normalize_equipment(system_elements_path)
    write_json(records, OUTPUT_JSON_PATH, system_elements_path)
    write_csv(records, OUTPUT_CSV_PATH)
    print(f"Wrote {len(records)} equipment records to {OUTPUT_JSON_PATH}")
    print(f"Wrote {len(records)} equipment records to {OUTPUT_CSV_PATH}")


if __name__ == "__main__":
    main()

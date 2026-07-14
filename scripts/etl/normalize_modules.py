"""Normalize module-to-equipment links from the module list workbook.

The module workbook is the PDM-centric starting point for the dashboard. This
step unpivots Equipment 1 through Equipment 9 into raw link records without
performing final matching against SystemElements.
"""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

try:
    from .file_discovery import get_input_files
    from .json_utils import file_metadata, write_json as write_json_payload
except ImportError:
    from file_discovery import get_input_files
    from json_utils import file_metadata, write_json as write_json_payload


PROJECT_ROOT = Path(__file__).resolve().parents[2]
OUTPUT_JSON_PATH = (
    PROJECT_ROOT / "frontend" / "public" / "data" / "module_equipment_links_raw.json"
)
OUTPUT_CSV_PATH = (
    PROJECT_ROOT / "frontend" / "public" / "data" / "module_equipment_links_raw.csv"
)

SOURCE_COLUMNS = [
    "PDM NAME",
    "MODULE TYPE: 14 TYPES",
    "LENGTH",
    "WIDTH",
    "HEIGHT",
    "WEIGHT",
    "Equipment 1",
    "Equipment 2",
    "Equipment 3",
    "Equipment 4",
    "Equipment 5",
    "Equipment 6",
    "Equipment 7",
    "Equipment 8",
    "Equipment 9",
]

EQUIPMENT_COLUMNS = [f"Equipment {index}" for index in range(1, 10)]

OUTPUT_FIELDS = [
    "pdm_name",
    "module_type",
    "length",
    "width",
    "height",
    "weight",
    "source_equipment_column",
    "source_equipment_label",
    "normalized_equipment_id",
    "matched_equipment_id",
    "match_status",
]


def is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")


def clean_text(value: Any) -> str | None:
    if is_blank(value):
        return None
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_header(value: str) -> str:
    return " ".join(value.strip().casefold().split())


def make_header_map(header_row: tuple[Any, ...]) -> dict[str, int]:
    header_map: dict[str, int] = {}
    for index, header in enumerate(header_row):
        cleaned_header = clean_text(header)
        if cleaned_header is not None:
            header_map[cleaned_header] = index
    return header_map


def find_source_header(header_map: dict[str, int], expected_header: str) -> str | None:
    expected_normalized = normalize_header(expected_header)

    for actual_header in sorted(header_map):
        if actual_header == expected_header:
            return actual_header

    for actual_header in sorted(header_map):
        if normalize_header(actual_header) == expected_normalized:
            return actual_header

    for actual_header in sorted(header_map):
        if normalize_header(actual_header).startswith(expected_normalized):
            return actual_header

    return None


def make_source_column_map(header_map: dict[str, int]) -> dict[str, int]:
    source_column_map: dict[str, int] = {}
    missing_columns: list[str] = []

    for source_column in SOURCE_COLUMNS:
        actual_header = find_source_header(header_map, source_column)
        if actual_header is None:
            missing_columns.append(source_column)
        else:
            source_column_map[source_column] = header_map[actual_header]

    if missing_columns:
        missing_display = ", ".join(missing_columns)
        raise ValueError(f"Missing expected module list columns: {missing_display}")

    return source_column_map


def get_cell(
    row: tuple[Any, ...],
    source_column_map: dict[str, int],
    source_column: str,
) -> Any:
    index = source_column_map[source_column]
    if index >= len(row):
        return None
    return row[index]


def normalize_equipment_id(source_equipment_label: str) -> str:
    if source_equipment_label.startswith("IAD06-"):
        return source_equipment_label
    return f"IAD06-{source_equipment_label}"


def make_link_record(
    row: tuple[Any, ...],
    source_column_map: dict[str, int],
    source_equipment_column: str,
    source_equipment_label: str,
) -> dict[str, Any]:
    return {
        "pdm_name": clean_text(get_cell(row, source_column_map, "PDM NAME")),
        "module_type": clean_text(
            get_cell(row, source_column_map, "MODULE TYPE: 14 TYPES")
        ),
        "length": clean_text(get_cell(row, source_column_map, "LENGTH")),
        "width": clean_text(get_cell(row, source_column_map, "WIDTH")),
        "height": clean_text(get_cell(row, source_column_map, "HEIGHT")),
        "weight": clean_text(get_cell(row, source_column_map, "WEIGHT")),
        "source_equipment_column": source_equipment_column,
        "source_equipment_label": source_equipment_label,
        "normalized_equipment_id": normalize_equipment_id(source_equipment_label),
        "matched_equipment_id": None,
        "match_status": "pending_match",
    }


def normalize_modules(module_list_path: str) -> list[dict]:
    workbook = load_workbook(module_list_path, read_only=True, data_only=True)
    try:
        sheet = workbook[workbook.sheetnames[0]]
        rows = sheet.iter_rows(values_only=True)
        try:
            header_row = next(rows)
        except StopIteration:
            return []

        header_map = make_header_map(header_row)
        source_column_map = make_source_column_map(header_map)

        records: list[dict] = []
        for row in rows:
            for equipment_column in EQUIPMENT_COLUMNS:
                source_equipment_label = clean_text(
                    get_cell(row, source_column_map, equipment_column)
                )
                if source_equipment_label is None:
                    continue
                records.append(
                    make_link_record(
                        row,
                        source_column_map,
                        equipment_column,
                        source_equipment_label,
                    )
                )
    finally:
        workbook.close()

    return records


def write_json(records: list[dict], output_path: Path, source_path: str) -> None:
    write_json_payload(
        output_path,
        {
            "source_file": file_metadata(source_path),
            "records": records,
        },
    )


def write_csv(records: list[dict], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8", newline="") as output_file:
        writer = csv.DictWriter(output_file, fieldnames=OUTPUT_FIELDS)
        writer.writeheader()
        writer.writerows(records)


def main() -> None:
    module_list_path = get_input_files()["module_list"]
    records = normalize_modules(module_list_path)
    write_json(records, OUTPUT_JSON_PATH, module_list_path)
    write_csv(records, OUTPUT_CSV_PATH)
    print(f"Wrote {len(records)} module equipment links to {OUTPUT_JSON_PATH}")
    print(f"Wrote {len(records)} module equipment links to {OUTPUT_CSV_PATH}")


if __name__ == "__main__":
    main()

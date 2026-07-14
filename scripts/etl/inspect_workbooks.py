"""Inspect raw Excel workbooks and write a JSON summary.

This script is intentionally inspection-only. It reads workbook shape,
headers, and sample records without transforming the source data into the
dashboard data model.
"""

from __future__ import annotations

import json
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

try:
    from .file_discovery import get_input_files
    from .json_utils import file_metadata, selected_input_files_metadata
except ImportError:
    from file_discovery import get_input_files
    from json_utils import file_metadata, selected_input_files_metadata


PROJECT_ROOT = Path(__file__).resolve().parents[2]
OUTPUT_PATH = PROJECT_ROOT / "frontend" / "public" / "data" / "workbook_inspection.json"

WORKBOOK_KEYS = ["system_elements", "module_list", "cases"]

EXPECTED_FIELDS_BY_WORKBOOK_KEY = {
    "module_list": [
        "PDM NAME",
        "MODULE TYPE: 14 TYPES",
        "LENGTH",
        "WIDTH",
        "HEIGHT",
        "WEIGHT",
        "Equipment 1",
        "Equipment 2",
        "Equipment 3",
        "Equipment 4",
        "Equipment 5",
        "Equipment 6",
        "Equipment 7",
        "Equipment 8",
        "Equipment 9",
    ],
    "system_elements": [
        "Unique ID",
        "Type",
        "Status",
        "Parent",
        "System",
        "Open Issues",
        "NETA Complete: Completed",
        "NETA Test Report",
        "Equipment Manufacturer",
        "Equipment Model",
        "Equipment Serial Number",
    ],
    "cases": [
        "Item #",
        "Category",
        "Status",
        "Priority",
        "Summary",
        "System Elements",
        "Issue Image",
        "Corrective Images",
        "Reported On",
        "Due",
        "Assigned to",
        "Created",
        "Last Updated",
    ],
}


def is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")


def last_non_blank_index(values: tuple[Any, ...]) -> int:
    for index in range(len(values) - 1, -1, -1):
        if not is_blank(values[index]):
            return index
    return -1


def make_json_value(value: Any) -> Any:
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    return value


def make_unique_headers(raw_headers: tuple[Any, ...], column_count: int) -> list[str]:
    headers: list[str] = []
    seen: dict[str, int] = {}

    for index in range(column_count):
        value = raw_headers[index] if index < len(raw_headers) else None
        header = str(value).strip() if not is_blank(value) else f"Unnamed: {index + 1}"

        count = seen.get(header, 0)
        seen[header] = count + 1
        if count:
            header = f"{header}_{count + 1}"

        headers.append(header)

    return headers


def normalize_header(value: str) -> str:
    return " ".join(value.split()).casefold()


def match_expected_fields(
    expected_fields: list[str],
    headers: set[str],
) -> list[dict[str, str | None]]:
    matches: list[dict[str, str | None]] = []

    for expected_field in expected_fields:
        expected_normalized = normalize_header(expected_field)
        match = {
            "expected": expected_field,
            "actual": None,
            "match_type": None,
        }

        for header in sorted(headers):
            if header == expected_field:
                match["actual"] = header
                match["match_type"] = "exact"
                break

        if match["actual"] is None:
            for header in sorted(headers):
                if normalize_header(header) == expected_normalized:
                    match["actual"] = header
                    match["match_type"] = "normalized"
                    break

        if match["actual"] is None:
            for header in sorted(headers):
                normalized_header = normalize_header(header)
                if normalized_header.startswith(expected_normalized):
                    match["actual"] = header
                    match["match_type"] = "prefix"
                    break

        matches.append(match)

    return matches


def row_to_record(row: tuple[Any, ...], headers: list[str]) -> dict[str, Any]:
    return {
        header: make_json_value(row[index]) if index < len(row) else None
        for index, header in enumerate(headers)
    }


def inspect_sheet(sheet: Any) -> dict[str, Any]:
    header_row: tuple[Any, ...] | None = None
    header_row_number: int | None = None
    row_count = 0
    column_count = 0
    sample_rows: list[tuple[Any, ...]] = []

    for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        row_width = last_non_blank_index(row) + 1
        if row_width == 0:
            continue

        if header_row is None:
            header_row = row
            header_row_number = row_number
            column_count = row_width
            continue

        row_count += 1
        column_count = max(column_count, row_width)
        if len(sample_rows) < 10:
            sample_rows.append(row)

    if header_row is None:
        return {
            "sheet_name": sheet.title,
            "header_row_number": None,
            "row_count": 0,
            "column_count": 0,
            "column_headers": [],
            "first_10_rows": [],
        }

    headers = make_unique_headers(header_row, column_count)

    return {
        "sheet_name": sheet.title,
        "header_row_number": header_row_number,
        "row_count": row_count,
        "column_count": column_count,
        "column_headers": headers,
        "first_10_rows": [row_to_record(row, headers) for row in sample_rows],
    }


def inspect_workbook(workbook_key: str, workbook_path: str) -> dict[str, Any]:
    workbook_path = Path(workbook_path)
    if not workbook_path.exists():
        raise FileNotFoundError(f"Missing workbook: {workbook_path}")

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        sheets = [inspect_sheet(workbook[sheet_name]) for sheet_name in workbook.sheetnames]
    finally:
        workbook.close()

    all_headers = {header for sheet in sheets for header in sheet["column_headers"]}
    expected_fields = EXPECTED_FIELDS_BY_WORKBOOK_KEY.get(workbook_key, [])
    expected_field_matches = match_expected_fields(expected_fields, all_headers)

    return {
        "workbook_key": workbook_key,
        "workbook_name": workbook_path.name,
        "path": str(workbook_path.relative_to(PROJECT_ROOT)),
        "source_file": file_metadata(workbook_path),
        "sheet_names": [sheet["sheet_name"] for sheet in sheets],
        "expected_fields": expected_fields,
        "expected_field_matches": expected_field_matches,
        "expected_fields_present": [
            match["expected"]
            for match in expected_field_matches
            if match["actual"] is not None
        ],
        "expected_fields_missing": [
            match["expected"]
            for match in expected_field_matches
            if match["actual"] is None
        ],
        "sheets": sheets,
    }


def main() -> None:
    input_files = get_input_files()
    inspection = {
        "generated_at": datetime.now().astimezone().isoformat(),
        "selected_input_files": selected_input_files_metadata(input_files),
        "project_context": {
            "dashboard_orientation": "PDM-centric",
            "main_view": (
                "Start from PDM Name, then show related equipment, issues, "
                "NETA status, and test reports."
            ),
            "scope": "Workbook inspection only; no dashboard transformation is applied.",
        },
        "workbooks": [
            inspect_workbook(workbook_key, input_files[workbook_key])
            for workbook_key in WORKBOOK_KEYS
        ],
    }

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(json.dumps(inspection, indent=2), encoding="utf-8")
    print(f"Wrote inspection JSON to {OUTPUT_PATH}")


if __name__ == "__main__":
    main()

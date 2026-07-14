"""Build EPS field-test execution data from the daily tracker workflow.

This output intentionally stays separate from the PDM/NETA/issue datasets. The
EPS tracker records field execution progress; Infralink NETA completion remains
the source of truth for final closeout.
"""

from __future__ import annotations

import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

try:
    from .json_utils import file_metadata, load_records_json, write_json
except ImportError:
    from json_utils import file_metadata, load_records_json, write_json


PROJECT_ROOT = Path(__file__).resolve().parents[2]
DATA_DIR = PROJECT_ROOT / "frontend" / "public" / "data"
EPS_ROOT = PROJECT_ROOT.parent / "IAD6_EPS_Testing_Tracker"
EPS_EXCEL_DIR = EPS_ROOT / "Excel"
EPS_TRACKER_PATH = EPS_EXCEL_DIR / "IAD06 TRACKER.xlsx"
EPS_TRACKER_SOURCE_DIR = EPS_EXCEL_DIR / "EPS_Trackers"
DAILY_TESTED_EQUIPMENT_PATH = EPS_ROOT / "daily_tested_equipment.md"

MODULE_LINKS_PATH = DATA_DIR / "module_equipment_links.json"
EQUIPMENT_PATH = DATA_DIR / "equipment.json"

SUMMARY_OUTPUT_PATH = DATA_DIR / "eps_test_summary.json"
PDM_OUTPUT_PATH = DATA_DIR / "eps_pdm_execution.json"
MODULE_OUTPUT_PATH = DATA_DIR / "eps_module_execution.json"
TEST_ITEMS_OUTPUT_PATH = DATA_DIR / "eps_test_items.json"
FAILED_OUTPUT_PATH = DATA_DIR / "eps_failed_items.json"
INCOMPLETE_OUTPUT_PATH = DATA_DIR / "eps_incomplete_items.json"
NOT_FOUND_OUTPUT_PATH = DATA_DIR / "eps_not_found_items.json"
SNAPSHOT_DIR = DATA_DIR / "eps_history" / "snapshots"

TRACKER_SHEET = "TRACKER"
ITEMS_NOT_IN_TRACKER_SHEET = "ITEMS NOT IN TRACKER TESTED"
LOOKBACK_DAYS = 7

STATUS_COMPLETE = "Complete"
STATUS_WAITING_INFRALINK_NETA = "Complete, Waiting Infralink NETA Completion"
STATUS_FAILED = "Failed"
ITEM_STATUS_PASSED = "Passed"
ITEM_STATUS_FIXED = "Fixed"
ITEM_STATUS_PASSED_NOT_IN_TRACKER = "Passed - Not In Tracker"
ITEM_STATUS_FIXED_NOT_IN_TRACKER = "Fixed - Not In Tracker"
STATUS_PARTIAL = "Partial"
STATUS_NOT_STARTED = "Not Started"
STATUS_NO_TRACKER_RECORDS = "No Tracker Records"

STATUS_ORDER = {
    STATUS_FAILED: 0,
    STATUS_PARTIAL: 1,
    STATUS_WAITING_INFRALINK_NETA: 2,
    STATUS_COMPLETE: 3,
    STATUS_NOT_STARTED: 4,
    STATUS_NO_TRACKER_RECORDS: 5,
}


@dataclass(frozen=True)
class TrackerRecord:
    row_number: int
    equipment_name: str
    equipment_key: str
    substation: str
    substation_key: str
    test_type: str
    equipment_type: str
    follow_up_req: str
    comments: str
    date_tested: str


@dataclass(frozen=True)
class NotFoundTestItem:
    status: str
    equipment_name: str
    alias_checked: str = ""
    retested_at: date | None = None


@dataclass(frozen=True)
class DailyTestedEquipmentHistory:
    source_file: Path
    records_by_date: dict[date, dict[str, set[str]]]

    @property
    def dates(self) -> list[date]:
        return sorted(self.records_by_date)

    @property
    def latest_date(self) -> date | None:
        dates = self.dates
        return dates[-1] if dates else None


@dataclass(frozen=True)
class ExecutionEquipmentSets:
    passed_input_equipment: set[str]
    failed_input_equipment: set[str]
    passed_equipment: set[str]
    failed_equipment: set[str]
    snapshot_passed_equipment: set[str]
    snapshot_failed_equipment: set[str]
    completed_equipment: set[str]
    alias_matches: dict[str, str]


def clean(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().split())


def key(value: Any) -> str:
    return clean(value).upper()


def compact_equipment_key(value: Any) -> str:
    normalized = key(value)
    if normalized.startswith("IAD06-"):
        return normalized.removeprefix("IAD06-")
    return normalized


def is_blank(value: Any) -> bool:
    return value is None or str(value).strip() == ""


def pdu_feeder_breaker_alias_key(equipment_key: str) -> str:
    match = re.fullmatch(r"(PDU6-[^-]+-[^-]+-FB)([1-8])", equipment_key)
    if not match:
        return ""

    number = int(match.group(2))
    if 1 <= number <= 4:
        return f"{match.group(1)}{number + 4}"
    return f"{match.group(1)}{number - 4}"


def tracker_key_for_equipment(
    equipment_key: str,
    tracker_equipment_keys: set[str],
) -> tuple[str, str]:
    if equipment_key in tracker_equipment_keys:
        return equipment_key, ""

    alias_key = pdu_feeder_breaker_alias_key(equipment_key)
    if alias_key and alias_key in tracker_equipment_keys:
        return alias_key, alias_key

    return "", alias_key


def tracker_match_keys(
    equipment_keys: set[str],
    tracker_equipment_keys: set[str],
) -> tuple[set[str], dict[str, str]]:
    matched_keys: set[str] = set()
    alias_matches: dict[str, str] = {}

    for equipment_key in equipment_keys:
        matched_key, alias_key = tracker_key_for_equipment(
            equipment_key, tracker_equipment_keys
        )
        if matched_key:
            matched_keys.add(matched_key)
        if alias_key and matched_key == alias_key:
            alias_matches[equipment_key] = alias_key

    return matched_keys, alias_matches


def unmatched_input_equipment_keys(
    equipment_keys: set[str],
    tracker_equipment_keys: set[str],
) -> set[str]:
    return {
        equipment_key
        for equipment_key in equipment_keys
        if not tracker_key_for_equipment(equipment_key, tracker_equipment_keys)[0]
    }


def build_not_found_test_items(
    passed_equipment: set[str],
    failed_equipment: set[str],
    tracker_equipment_keys: set[str],
    retested_dates: dict[str, date] | None = None,
) -> list[NotFoundTestItem]:
    retested_dates = retested_dates or {}
    items: list[NotFoundTestItem] = []
    for status, equipment_keys in [
        ("Passed", passed_equipment),
        ("Failed", failed_equipment),
    ]:
        for equipment_key in sorted(equipment_keys):
            matched_key, alias_checked = tracker_key_for_equipment(
                equipment_key,
                tracker_equipment_keys,
            )
            if matched_key:
                continue
            items.append(
                NotFoundTestItem(
                    status=status,
                    equipment_name=equipment_key,
                    alias_checked=alias_checked,
                    retested_at=(
                        retested_dates.get(equipment_key) if status == "Passed" else None
                    ),
                )
            )
    return items


def build_execution_equipment_sets(
    passed_input_equipment: set[str],
    failed_input_equipment: set[str],
    tracker_equipment_keys: set[str],
) -> ExecutionEquipmentSets:
    passed_equipment, passed_alias_matches = tracker_match_keys(
        passed_input_equipment,
        tracker_equipment_keys,
    )
    failed_equipment, failed_alias_matches = tracker_match_keys(
        failed_input_equipment,
        tracker_equipment_keys,
    )
    failed_equipment -= passed_equipment
    snapshot_passed_equipment = passed_equipment | unmatched_input_equipment_keys(
        passed_input_equipment,
        tracker_equipment_keys,
    )
    snapshot_failed_equipment = failed_equipment | unmatched_input_equipment_keys(
        failed_input_equipment,
        tracker_equipment_keys,
    )

    return ExecutionEquipmentSets(
        passed_input_equipment=passed_input_equipment,
        failed_input_equipment=failed_input_equipment,
        passed_equipment=passed_equipment,
        failed_equipment=failed_equipment,
        snapshot_passed_equipment=snapshot_passed_equipment,
        snapshot_failed_equipment=snapshot_failed_equipment,
        completed_equipment=passed_equipment | failed_equipment,
        alias_matches={**passed_alias_matches, **failed_alias_matches},
    )


def strip_markdown_bullet(line: str) -> str:
    while line.startswith("-"):
        line = line[1:].strip()
    return line


def clean_daily_equipment_name(line: str) -> str | None:
    line = strip_markdown_bullet(line.strip())
    if not line:
        return None

    ignored_words = {"equipment name", "type", "equipment type", "failed", "tested"}
    if line.lower() in ignored_words or line.startswith("#"):
        return None

    if " " in line or "-" not in line or not any(char.isdigit() for char in line):
        return None

    return line


def parse_equipment_sections(path: Path) -> tuple[set[str], set[str]]:
    tested: set[str] = set()
    failed: set[str] = set()
    section = "tested"

    if not path.exists():
        return tested, failed

    for raw_line in path.read_text(encoding="utf-8-sig").splitlines():
        line = raw_line.strip()
        if not line:
            continue

        if line.startswith("#"):
            heading = key(line.lstrip("#").strip())
            if "FAILED EQUIPMENT" in heading or heading == "FAILED":
                section = "failed"
            elif "TESTED EQUIPMENT" in heading or heading == "TESTED":
                section = "tested"
            continue

        item = strip_markdown_bullet(line)
        if item:
            normalized = compact_equipment_key(item)
            if section == "failed":
                failed.add(normalized)
            else:
                tested.add(normalized)

    return tested - failed, failed


def parse_daily_equipment_date(value: str, reference_date: date) -> date | None:
    text = clean(value).lstrip("#").strip()
    if not text:
        return None

    for pattern in ("%Y-%m-%d", "%m-%d", "%m/%d"):
        try:
            parsed = datetime.strptime(text, pattern).date()
        except ValueError:
            continue

        if pattern == "%Y-%m-%d":
            return parsed
        return date(reference_date.year, parsed.month, parsed.day)

    match = re.fullmatch(r"(\d{1,2})[-/](\d{1,2})", text)
    if not match:
        return None
    return date(reference_date.year, int(match.group(1)), int(match.group(2)))


def parse_daily_tested_equipment(
    path: Path,
    reference_date: date | None = None,
) -> DailyTestedEquipmentHistory:
    reference_date = reference_date or snapshot_date_today()
    records_by_date: dict[date, dict[str, set[str]]] = defaultdict(
        lambda: {"tested": set(), "retested": set(), "failed": set()}
    )
    section = ""
    current_date: date | None = None

    if not path.exists():
        return DailyTestedEquipmentHistory(path, {})

    for raw_line in path.read_text(encoding="utf-8-sig").splitlines():
        line = raw_line.strip()
        if not line:
            continue

        if line.startswith("##"):
            parsed_date = parse_daily_equipment_date(
                line.lstrip("#").strip(),
                reference_date,
            )
            current_date = parsed_date
            if current_date:
                _ = records_by_date[current_date]
            continue

        if line.startswith("#"):
            heading = key(line.lstrip("#").strip())
            if heading in {"RETESTED AND PASSED", "RETEST AND PASSED"}:
                section = "retested"
            elif "FAILED EQUIPMENT" in heading or heading == "FAILED":
                section = "failed"
            elif "TESTED EQUIPMENT" in heading or heading == "TESTED":
                section = "tested"
            continue

        if current_date is None or section not in {"tested", "retested", "failed"}:
            continue

        item = clean_daily_equipment_name(line)
        if not item:
            continue
        records_by_date[current_date][section].add(compact_equipment_key(item))

    return DailyTestedEquipmentHistory(
        source_file=path,
        records_by_date={
            daily_date: dict(records) for daily_date, records in records_by_date.items()
        },
    )


def daily_equipment_on_date(
    history: DailyTestedEquipmentHistory,
    source_date: date,
) -> tuple[set[str], set[str]]:
    records = history.records_by_date.get(source_date) or {}
    passed = set(records.get("tested") or set()) | set(
        records.get("retested") or set()
    )
    failed = set(records.get("failed") or set()) - passed
    return passed, failed


def retested_equipment_on_date(
    history: DailyTestedEquipmentHistory,
    source_date: date,
) -> set[str]:
    records = history.records_by_date.get(source_date) or {}
    return set(records.get("retested") or set())


def cumulative_retested_equipment_dates(
    history: DailyTestedEquipmentHistory,
    through_date: date,
) -> dict[str, date]:
    retested_dates: dict[str, date] = {}
    for daily_date in history.dates:
        if daily_date > through_date:
            break
        for equipment_key in retested_equipment_on_date(history, daily_date):
            retested_dates[equipment_key] = daily_date
    return retested_dates


def current_retested_equipment_dates(
    history: DailyTestedEquipmentHistory,
    through_date: date,
    passed_equipment: set[str],
) -> dict[str, date]:
    return {
        equipment_key: retested_at
        for equipment_key, retested_at in cumulative_retested_equipment_dates(
            history, through_date
        ).items()
        if equipment_key in passed_equipment
    }


def tracker_retested_equipment_dates(
    retested_input_dates: dict[str, date],
    tracker_equipment_keys: set[str],
) -> dict[str, date]:
    matched_dates: dict[str, date] = {}
    for equipment_key, retested_at in retested_input_dates.items():
        matched_key, _alias_key = tracker_key_for_equipment(
            equipment_key, tracker_equipment_keys
        )
        if matched_key:
            matched_dates[matched_key] = retested_at
    return matched_dates


def cumulative_daily_equipment(
    history: DailyTestedEquipmentHistory,
    through_date: date,
) -> tuple[set[str], set[str]]:
    passed: set[str] = set()
    failed: set[str] = set()

    for daily_date in history.dates:
        if daily_date > through_date:
            break

        tested_today, failed_today = daily_equipment_on_date(history, daily_date)
        for equipment_key in tested_today:
            failed.discard(equipment_key)
            passed.add(equipment_key)
        for equipment_key in failed_today:
            passed.discard(equipment_key)
            failed.add(equipment_key)

    return passed, failed


def choose_baseline_daily_date(
    dates: Iterable[date],
    current_source_date: date,
    lookback_days: int,
    tolerance_days: int = 2,
) -> date | None:
    target_date = current_source_date - timedelta(days=lookback_days)
    sorted_dates = sorted(
        daily_date for daily_date in dates if daily_date < current_source_date
    )

    older_or_target = [
        daily_date for daily_date in sorted_dates if daily_date <= target_date
    ]
    if older_or_target:
        return older_or_target[-1]

    min_age_days = max(1, lookback_days - tolerance_days)
    max_age_days = lookback_days + tolerance_days
    fallback_candidates = [
        daily_date
        for daily_date in sorted_dates
        if min_age_days <= (current_source_date - daily_date).days <= max_age_days
    ]
    if not fallback_candidates:
        return None

    return min(
        fallback_candidates,
        key=lambda daily_date: abs((daily_date - target_date).days),
    )


def daily_activity_summary(
    history: DailyTestedEquipmentHistory,
    source_date: date | None,
    tracker_equipment_keys: set[str],
) -> dict[str, Any]:
    if source_date is None:
        return {
            "available": False,
            "source": "daily_tested_equipment",
            "source_file": (
                file_metadata(history.source_file)
                if history.source_file.exists()
                else None
            ),
            "source_date_label": None,
            "new_tested_count": 0,
            "new_failed_count": 0,
            "repaired_count": 0,
            "new_tested_equipment": [],
            "new_failed_equipment": [],
            "repaired_equipment": [],
        }

    tested, failed = daily_equipment_on_date(history, source_date)
    retested = retested_equipment_on_date(history, source_date)
    equipment_sets = build_execution_equipment_sets(tested, failed, tracker_equipment_keys)
    retested_sets = build_execution_equipment_sets(
        retested, set(), tracker_equipment_keys
    )
    return {
        "available": True,
        "source": "daily_tested_equipment",
        "source_file": file_metadata(history.source_file),
        "source_date_label": source_date.isoformat(),
        "new_tested_count": len(equipment_sets.snapshot_passed_equipment),
        "new_failed_count": len(equipment_sets.snapshot_failed_equipment),
        "repaired_count": len(retested_sets.snapshot_passed_equipment),
        "new_tested_equipment": sorted(equipment_sets.snapshot_passed_equipment),
        "new_failed_equipment": sorted(equipment_sets.snapshot_failed_equipment),
        "repaired_equipment": sorted(retested_sets.snapshot_passed_equipment),
    }


def build_daily_history_summary(
    history: DailyTestedEquipmentHistory,
    tracker_equipment_keys: set[str],
) -> dict[str, Any]:
    entries: list[dict[str, Any]] = []

    for source_date in history.dates:
        daily_tested, daily_failed = daily_equipment_on_date(history, source_date)
        daily_retested = retested_equipment_on_date(history, source_date)
        daily_sets = build_execution_equipment_sets(
            daily_tested,
            daily_failed,
            tracker_equipment_keys,
        )
        cumulative_tested, cumulative_failed = cumulative_daily_equipment(
            history,
            source_date,
        )
        cumulative_sets = build_execution_equipment_sets(
            cumulative_tested,
            cumulative_failed,
            tracker_equipment_keys,
        )
        daily_retested_sets = build_execution_equipment_sets(
            daily_retested,
            set(),
            tracker_equipment_keys,
        )
        cumulative_retested = current_retested_equipment_dates(
            history,
            source_date,
            cumulative_tested,
        )
        cumulative_retested_sets = build_execution_equipment_sets(
            set(cumulative_retested),
            set(),
            tracker_equipment_keys,
        )

        entries.append(
            {
                "date": source_date.isoformat(),
                "daily_passed_count": len(daily_sets.snapshot_passed_equipment),
                "daily_failed_count": len(daily_sets.snapshot_failed_equipment),
                "daily_passed_equipment": sorted(daily_sets.snapshot_passed_equipment),
                "daily_failed_equipment": sorted(daily_sets.snapshot_failed_equipment),
                "daily_fixed_count": len(
                    daily_retested_sets.snapshot_passed_equipment
                ),
                "daily_fixed_equipment": sorted(
                    daily_retested_sets.snapshot_passed_equipment
                ),
                "cumulative_passed_count": len(cumulative_sets.snapshot_passed_equipment),
                "cumulative_failed_count": len(cumulative_sets.snapshot_failed_equipment),
                "cumulative_passed_equipment": sorted(
                    cumulative_sets.snapshot_passed_equipment
                ),
                "cumulative_failed_equipment": sorted(
                    cumulative_sets.snapshot_failed_equipment
                ),
                "cumulative_fixed_count": len(
                    cumulative_retested_sets.snapshot_passed_equipment
                ),
                "cumulative_fixed_equipment": sorted(
                    cumulative_retested_sets.snapshot_passed_equipment
                ),
            }
        )

    latest_date = history.latest_date
    baseline_date = (
        choose_baseline_daily_date(history.dates, latest_date, LOOKBACK_DAYS)
        if latest_date
        else None
    )

    return {
        "available": bool(entries),
        "source": "daily_tested_equipment",
        "latest_date": latest_date.isoformat() if latest_date else None,
        "default_current_date": latest_date.isoformat() if latest_date else None,
        "default_baseline_date": baseline_date.isoformat() if baseline_date else None,
        "dates": [entry["date"] for entry in entries],
        "entries": entries,
    }


def worksheet_table_bounds(ws) -> tuple[int, int, int, int]:
    if ws.tables:
        first_table_name = next(iter(ws.tables.keys()))
        return range_boundaries(ws.tables[first_table_name].ref)
    return 1, 1, ws.max_column, ws.max_row


def header_map(ws, header_row: int, min_col: int, max_col: int) -> dict[str, int]:
    headers: dict[str, int] = {}
    for col in range(min_col, max_col + 1):
        header = key(ws.cell(header_row, col).value)
        if header:
            headers[header] = col
    return headers


def worksheet_by_name(workbook, sheet_name: str):
    if sheet_name in workbook.sheetnames:
        return workbook[sheet_name]
    raise ValueError(f"Workbook is missing required sheet: {sheet_name}")


def load_tracker_records(path: Path) -> tuple[dict[str, list[TrackerRecord]], set[str]]:
    workbook = load_workbook(path, data_only=True)
    worksheet = worksheet_by_name(workbook, TRACKER_SHEET)
    min_col, header_row, max_col, max_row = worksheet_table_bounds(worksheet)
    headers = header_map(worksheet, header_row, min_col, max_col)

    required_headers = ["EQUIPMENT NAME", "SUBSTATION"]
    missing_headers = [name for name in required_headers if name not in headers]
    if missing_headers:
        raise ValueError(
            f"Tracker is missing required columns: {', '.join(missing_headers)}"
        )

    equipment_col = headers["EQUIPMENT NAME"]
    substation_col = headers["SUBSTATION"]
    type_col = headers.get("TYPE")
    equipment_type_col = headers.get("EQUIPMENT TYPE")
    follow_up_req_col = headers.get("FOLLOW UP REQ")
    comments_col = headers.get("COMMENTS")
    date_tested_col = headers.get("DATE TESTED")

    by_substation: dict[str, list[TrackerRecord]] = defaultdict(list)
    equipment_keys: set[str] = set()
    for row_number in range(header_row + 1, max_row + 1):
        equipment_name = clean(worksheet.cell(row_number, equipment_col).value)
        substation = clean(worksheet.cell(row_number, substation_col).value)
        if not equipment_name or not substation:
            continue

        record = TrackerRecord(
            row_number=row_number,
            equipment_name=equipment_name,
            equipment_key=compact_equipment_key(equipment_name),
            substation=substation,
            substation_key=compact_equipment_key(substation),
            test_type=(
                clean(worksheet.cell(row_number, type_col).value) if type_col else ""
            ),
            equipment_type=(
                clean(worksheet.cell(row_number, equipment_type_col).value)
                if equipment_type_col
                else ""
            ),
            follow_up_req=(
                clean(worksheet.cell(row_number, follow_up_req_col).value)
                if follow_up_req_col
                else ""
            ),
            comments=(
                clean(worksheet.cell(row_number, comments_col).value)
                if comments_col
                else ""
            ),
            date_tested=(
                clean(worksheet.cell(row_number, date_tested_col).value)
                if date_tested_col
                else ""
            ),
        )
        by_substation[record.substation_key].append(record)
        equipment_keys.add(record.equipment_key)

    return dict(by_substation), equipment_keys


def load_equipment_info() -> dict[str, dict[str, Any]]:
    records = load_records_json(EQUIPMENT_PATH)
    info_by_key: dict[str, dict[str, Any]] = {}

    for record in records:
        equipment_id = record.get("equipment_id")
        for lookup_key in {key(equipment_id), compact_equipment_key(equipment_id)}:
            if lookup_key:
                info_by_key[lookup_key] = record

    return info_by_key


def lookup_equipment_info(
    module_link: dict[str, Any],
    info_by_key: dict[str, dict[str, Any]],
) -> dict[str, Any]:
    candidates = [
        module_link.get("matched_equipment_id"),
        module_link.get("normalized_equipment_id"),
        module_link.get("source_equipment_label"),
    ]
    for candidate in candidates:
        for lookup_key in {key(candidate), compact_equipment_key(candidate)}:
            if lookup_key and lookup_key in info_by_key:
                return info_by_key[lookup_key]

    source_key = compact_equipment_key(module_link.get("source_equipment_label"))
    if source_key:
        suffix_matches = [
            (candidate_key, info)
            for candidate_key, info in info_by_key.items()
            if candidate_key.endswith(f"-{source_key}")
        ]
        if suffix_matches:
            return min(suffix_matches, key=lambda item: len(item[0]))[1]

    return {}


def is_pdm_name(value: Any) -> bool:
    return isinstance(value, str) and value.strip().casefold().startswith(
        ("iad06-pdm-", "iad6-pdm-")
    )


def get_effective_pdm_name(
    module_link: dict[str, Any], equipment_info: dict[str, Any]
) -> str | None:
    parent = equipment_info.get("parent")
    if is_pdm_name(parent):
        return str(parent).strip()

    pdm_name = module_link.get("pdm_name")
    if isinstance(pdm_name, str) and pdm_name.strip():
        return pdm_name.strip()
    return None


def link_with_effective_pdm(
    module_link: dict[str, Any],
    equipment_info: dict[str, Any],
) -> dict[str, Any]:
    effective_pdm_name = get_effective_pdm_name(module_link, equipment_info)
    if effective_pdm_name == module_link.get("pdm_name"):
        return module_link
    return {**module_link, "pdm_name": effective_pdm_name}


def get_neta_complete(info: dict[str, Any]) -> bool:
    value = info.get("neta_complete")
    if value is True:
        return True
    if isinstance(value, str):
        return key(value) not in {
            "",
            "N/A",
            "NA",
            "NOT COMPLETED",
            "NOT COMPLETE",
            "FALSE",
            "NO",
        }
    return False


def load_module_links() -> list[dict[str, Any]]:
    links = load_records_json(MODULE_LINKS_PATH)
    return [
        link
        for link in links
        if not is_blank(link.get("pdm_name"))
        and not is_blank(link.get("source_equipment_label"))
    ]


def module_link_lookup_keys(module_link: dict[str, Any]) -> set[str]:
    return {
        lookup_key
        for candidate in [
            module_link.get("source_equipment_label"),
            module_link.get("normalized_equipment_id"),
            module_link.get("matched_equipment_id"),
        ]
        for lookup_key in {key(candidate), compact_equipment_key(candidate)}
        if lookup_key
    }


def build_module_link_index(
    module_links: list[dict[str, Any]],
) -> dict[str, list[dict[str, Any]]]:
    index: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for module_link in module_links:
        for lookup_key in module_link_lookup_keys(module_link):
            index[lookup_key].append(module_link)
    return dict(index)


def find_module_link_for_equipment_key(
    equipment_key: str,
    module_link_index: dict[str, list[dict[str, Any]]],
    module_keys: set[str],
    info_by_key: dict[str, dict[str, Any]],
) -> tuple[dict[str, Any] | None, str]:
    equipment_info = info_by_key.get(equipment_key, {})
    candidates = [
        equipment_key,
        compact_equipment_key(equipment_info.get("equipment_id")),
        compact_equipment_key(equipment_info.get("parent")),
        prefix_module_key(equipment_key, module_keys),
        suffix_module_key(equipment_key, module_keys),
    ]

    seen: set[str] = set()
    for candidate in candidates:
        if not candidate or candidate in seen:
            continue
        seen.add(candidate)
        matches = module_link_index.get(candidate, [])
        if len(matches) == 1:
            return matches[0], "matched_module_link"
        if len(matches) > 1:
            return None, "ambiguous_module_link"

    return None, "unmatched_module_link"


def prefix_module_key(record_key: str, module_keys: set[str]) -> str:
    matches = [
        module_key
        for module_key in module_keys
        if record_key != module_key and record_key.startswith(f"{module_key}-")
    ]
    if not matches:
        return ""
    return max(matches, key=len)


def suffix_module_key(record_key: str, module_keys: set[str]) -> str:
    matches = [
        module_key
        for module_key in module_keys
        if record_key != module_key and record_key.endswith(f"-{module_key}")
    ]
    if not matches:
        return ""
    return max(matches, key=len)


def tracker_records_by_module(
    tracker_by_substation: dict[str, list[TrackerRecord]],
    module_keys: set[str],
) -> dict[str, list[TrackerRecord]]:
    records_by_module = {module_key: [] for module_key in module_keys}

    for records in tracker_by_substation.values():
        for record in records:
            matched_module_key = tracker_module_key_for_record(record, module_keys)
            if not matched_module_key:
                continue

            records_by_module.setdefault(matched_module_key, []).append(record)

    return records_by_module


def date_tested_indicates_tested(value: str) -> bool:
    normalized = key(value)
    if not normalized or normalized in {"N/A", "NA", "N/T", "NT", "NOT TESTED"}:
        return False
    return True


def record_has_follow_up_failure(record: TrackerRecord) -> bool:
    return bool(
        record.follow_up_req
        and record.comments
        and date_tested_indicates_tested(record.date_tested)
    )


def record_is_closed_without_markdown(record: TrackerRecord) -> bool:
    return bool(record.comments and date_tested_indicates_tested(record.date_tested))


def record_is_failed(
    record: TrackerRecord,
    failed_equipment: set[str],
    fixed_equipment: set[str] | None = None,
    passed_equipment: set[str] | None = None,
) -> bool:
    if fixed_equipment and record.equipment_key in fixed_equipment:
        return False
    if record.equipment_key in failed_equipment:
        return True
    if passed_equipment and record.equipment_key in passed_equipment:
        return False
    return record_has_follow_up_failure(record)


def record_is_complete(record: TrackerRecord, completed_equipment: set[str]) -> bool:
    return (
        record.equipment_key in completed_equipment
        or record_is_closed_without_markdown(record)
    )


def retest_comment(retested_at: date) -> str:
    return f"Retested and passed on {retested_at.isoformat()}."


def append_test_item_comment(existing: str, new_comment: str) -> str:
    existing = clean(existing)
    if not existing:
        return new_comment
    if key(new_comment) in key(existing):
        return existing
    return f"{existing}; {new_comment}"


def summarize_module_status(
    records: list[TrackerRecord],
    equipment_info: dict[str, Any],
    completed_equipment: set[str],
    failed_equipment: set[str],
    fixed_equipment: set[str] | None = None,
) -> tuple[str, list[TrackerRecord], list[TrackerRecord], list[TrackerRecord]]:
    if not records:
        return STATUS_NO_TRACKER_RECORDS, [], [], []

    passed_equipment = completed_equipment - failed_equipment
    completed_records = [
        record for record in records if record_is_complete(record, completed_equipment)
    ]
    failed_records = [
        record
        for record in records
        if record_is_failed(
            record,
            failed_equipment,
            fixed_equipment,
            passed_equipment,
        )
    ]
    incomplete_records = [
        record
        for record in records
        if not record_is_complete(record, completed_equipment)
    ]

    if failed_records:
        return STATUS_FAILED, completed_records, failed_records, incomplete_records

    if len(completed_records) == len(records):
        if get_neta_complete(equipment_info):
            return (
                STATUS_COMPLETE,
                completed_records,
                failed_records,
                incomplete_records,
            )
        return (
            STATUS_WAITING_INFRALINK_NETA,
            completed_records,
            failed_records,
            incomplete_records,
        )

    if not completed_records:
        return STATUS_NOT_STARTED, completed_records, failed_records, incomplete_records

    return STATUS_PARTIAL, completed_records, failed_records, incomplete_records


def item_record(
    *,
    module_link: dict[str, Any],
    equipment_info: dict[str, Any],
    tracker_record: TrackerRecord,
    item_status: str,
    retested_at: date | None = None,
) -> dict[str, Any]:
    comments = tracker_record.comments
    if retested_at:
        comments = append_test_item_comment(comments, retest_comment(retested_at))
    return {
        "item_status": item_status,
        "pdm_name": module_link.get("pdm_name"),
        "module_equipment": module_link.get("source_equipment_label"),
        "module_equipment_key": compact_equipment_key(
            module_link.get("source_equipment_label")
        ),
        "matched_equipment_id": module_link.get("matched_equipment_id"),
        "equipment_name": tracker_record.equipment_name,
        "equipment_key": tracker_record.equipment_key,
        "tracker_row": tracker_record.row_number,
        "tracker_type": tracker_record.test_type,
        "tracker_equipment_type": tracker_record.equipment_type,
        "follow_up_req": tracker_record.follow_up_req,
        "comments": comments,
        "date_tested": tracker_record.date_tested,
        "retested_and_passed": retested_at is not None,
        "retested_at": retested_at.isoformat() if retested_at else None,
        "equipment_serial_number": equipment_info.get("serial_number"),
        "equipment_manufacturer": equipment_info.get("manufacturer"),
        "equipment_model": equipment_info.get("model"),
        "neta_complete": equipment_info.get("neta_complete"),
        "neta_test_report": equipment_info.get("neta_test_report"),
    }


def infer_missing_tracker_type(equipment_name: str) -> str | None:
    normalized = key(equipment_name)
    if re.search(r"(?:^|[^A-Z0-9])CT(?:\d+)?(?:[^A-Z0-9]|$)", normalized):
        return "CT"
    return None


def not_found_item_record(
    *,
    item: NotFoundTestItem,
    module_link: dict[str, Any] | None,
    equipment_info: dict[str, Any],
    module_match_status: str,
) -> dict[str, Any]:
    equipment_key = compact_equipment_key(item.equipment_name)
    module_equipment = (
        module_link.get("source_equipment_label") if module_link else None
    )
    module_equipment_key = (
        compact_equipment_key(module_equipment) if module_equipment else equipment_key
    )
    if item.retested_at:
        item_status = ITEM_STATUS_FIXED_NOT_IN_TRACKER
    elif item.status == "Passed":
        item_status = ITEM_STATUS_PASSED_NOT_IN_TRACKER
    else:
        item_status = "Failed - Not In Tracker"
    inferred_tracker_type = infer_missing_tracker_type(item.equipment_name)
    comments = "Listed in daily report but can't found in the EPS tracker."
    if item.retested_at:
        comments = append_test_item_comment(
            retest_comment(item.retested_at), comments
        )

    return {
        "item_status": item_status,
        "source_status": item.status,
        "pdm_name": module_link.get("pdm_name") if module_link else None,
        "module_equipment": module_equipment,
        "module_equipment_key": module_equipment_key,
        "matched_equipment_id": (
            module_link.get("matched_equipment_id")
            if module_link
            else equipment_info.get("equipment_id")
        ),
        "equipment_name": item.equipment_name,
        "equipment_key": equipment_key,
        "tracker_row": None,
        "tracker_type": inferred_tracker_type,
        "tracker_equipment_type": inferred_tracker_type,
        "follow_up_req": "",
        "comments": comments,
        "date_tested": "",
        "retested_and_passed": item.retested_at is not None,
        "retested_at": item.retested_at.isoformat() if item.retested_at else None,
        "equipment_serial_number": equipment_info.get("serial_number"),
        "equipment_manufacturer": equipment_info.get("manufacturer"),
        "equipment_model": equipment_info.get("model"),
        "neta_complete": equipment_info.get("neta_complete"),
        "neta_test_report": equipment_info.get("neta_test_report"),
        "status": item.status,
        "alias_checked": item.alias_checked,
        "module_match_status": module_match_status,
        "reason": "No matching Equipment Name found in tracker.",
    }


def build_not_found_records(
    not_found_items: list[NotFoundTestItem],
    module_links: list[dict[str, Any]],
    module_keys: set[str],
    info_by_key: dict[str, dict[str, Any]],
) -> list[dict[str, Any]]:
    module_link_index = build_module_link_index(module_links)
    records: list[dict[str, Any]] = []

    for item in not_found_items:
        equipment_key = compact_equipment_key(item.equipment_name)
        module_link, module_match_status = find_module_link_for_equipment_key(
            equipment_key,
            module_link_index,
            module_keys,
            info_by_key,
        )
        equipment_info = info_by_key.get(equipment_key, {})
        if module_link:
            equipment_info = (
                lookup_equipment_info(module_link, info_by_key) or equipment_info
            )
            module_link = link_with_effective_pdm(module_link, equipment_info)
        records.append(
            not_found_item_record(
                item=item,
                module_link=module_link,
                equipment_info=equipment_info,
                module_match_status=module_match_status,
            )
        )

    return records


def flatten_tracker_records(
    tracker_by_substation: dict[str, list[TrackerRecord]],
) -> list[TrackerRecord]:
    return sorted(
        [record for records in tracker_by_substation.values() for record in records],
        key=lambda record: (
            record.substation_key,
            record.row_number,
            record.equipment_key,
        ),
    )


def find_module_link_for_tracker_record(
    record: TrackerRecord,
    module_link_index: dict[str, list[dict[str, Any]]],
    module_keys: set[str],
) -> tuple[dict[str, Any] | None, str]:
    candidates = [
        record.equipment_key if record.equipment_key in module_keys else "",
        prefix_module_key(record.equipment_key, module_keys),
        record.substation_key,
        prefix_module_key(record.substation_key, module_keys),
        suffix_module_key(record.substation_key, module_keys),
        suffix_module_key(record.equipment_key, module_keys),
    ]

    seen: set[str] = set()
    for candidate in candidates:
        if not candidate or candidate in seen:
            continue
        seen.add(candidate)
        matches = module_link_index.get(candidate, [])
        if len(matches) == 1:
            return matches[0], "matched_module_link"
        if len(matches) > 1:
            return None, "ambiguous_module_link"

    return None, "unmatched_module_link"


def tracker_module_key_for_record(record: TrackerRecord, module_keys: set[str]) -> str:
    equipment_candidates = [
        record.equipment_key if record.equipment_key in module_keys else "",
        prefix_module_key(record.equipment_key, module_keys),
    ]
    for candidate in equipment_candidates:
        if candidate:
            return candidate

    if record.substation_key in module_keys:
        return record.substation_key

    return suffix_module_key(record.substation_key, module_keys)


def tracker_item_record(
    *,
    tracker_record: TrackerRecord,
    module_link: dict[str, Any] | None,
    equipment_info: dict[str, Any],
    item_status: str,
    module_match_status: str,
    retested_at: date | None = None,
) -> dict[str, Any]:
    comments = tracker_record.comments
    if retested_at:
        comments = append_test_item_comment(comments, retest_comment(retested_at))
    return {
        "item_status": item_status,
        "pdm_name": (
            module_link.get("pdm_name")
            if module_link
            else (
                equipment_info.get("parent")
                if is_pdm_name(equipment_info.get("parent"))
                else None
            )
        ),
        "module_equipment": (
            module_link.get("source_equipment_label")
            if module_link
            else tracker_record.substation
        ),
        "module_equipment_key": (
            compact_equipment_key(module_link.get("source_equipment_label"))
            if module_link
            else tracker_record.substation_key
        ),
        "matched_equipment_id": (
            module_link.get("matched_equipment_id")
            if module_link
            else equipment_info.get("equipment_id")
        ),
        "equipment_name": tracker_record.equipment_name,
        "equipment_key": tracker_record.equipment_key,
        "tracker_row": tracker_record.row_number,
        "tracker_type": tracker_record.test_type,
        "tracker_equipment_type": tracker_record.equipment_type,
        "follow_up_req": tracker_record.follow_up_req,
        "comments": comments,
        "date_tested": tracker_record.date_tested,
        "retested_and_passed": retested_at is not None,
        "retested_at": retested_at.isoformat() if retested_at else None,
        "equipment_serial_number": equipment_info.get("serial_number"),
        "equipment_manufacturer": equipment_info.get("manufacturer"),
        "equipment_model": equipment_info.get("model"),
        "neta_complete": equipment_info.get("neta_complete"),
        "neta_test_report": equipment_info.get("neta_test_report"),
        "module_match_status": module_match_status,
    }


def build_tracker_test_items(
    tracker_by_substation: dict[str, list[TrackerRecord]],
    module_links: list[dict[str, Any]],
    module_keys: set[str],
    info_by_key: dict[str, dict[str, Any]],
    completed_equipment: set[str],
    failed_equipment: set[str],
    fixed_equipment_dates: dict[str, date] | None = None,
) -> list[dict[str, Any]]:
    fixed_equipment_dates = fixed_equipment_dates or {}
    fixed_equipment = set(fixed_equipment_dates)
    passed_equipment = completed_equipment - failed_equipment
    module_link_index = build_module_link_index(module_links)
    records: list[dict[str, Any]] = []

    for tracker_record in flatten_tracker_records(tracker_by_substation):
        module_link, module_match_status = find_module_link_for_tracker_record(
            tracker_record,
            module_link_index,
            module_keys,
        )
        equipment_info = info_by_key.get(tracker_record.equipment_key, {})
        if module_link:
            module_equipment_info = lookup_equipment_info(module_link, info_by_key)
            equipment_info = module_equipment_info or equipment_info
            module_link = link_with_effective_pdm(module_link, equipment_info)

        retested_at = fixed_equipment_dates.get(tracker_record.equipment_key)
        if record_is_failed(
            tracker_record,
            failed_equipment,
            fixed_equipment,
            passed_equipment,
        ):
            item_status = STATUS_FAILED
        elif record_is_complete(tracker_record, completed_equipment):
            item_status = ITEM_STATUS_FIXED if retested_at else ITEM_STATUS_PASSED
        else:
            item_status = "Not Tested"

        records.append(
            tracker_item_record(
                tracker_record=tracker_record,
                module_link=module_link,
                equipment_info=equipment_info,
                item_status=item_status,
                module_match_status=module_match_status,
                retested_at=retested_at,
            )
        )

    return records


def build_module_execution_records(
    module_links: list[dict[str, Any]],
    info_by_key: dict[str, dict[str, Any]],
    records_by_module: dict[str, list[TrackerRecord]],
    completed_equipment: set[str],
    failed_equipment: set[str],
    fixed_equipment_dates: dict[str, date] | None = None,
) -> tuple[
    list[dict[str, Any]],
    list[dict[str, Any]],
    list[dict[str, Any]],
    list[dict[str, Any]],
    set[str],
]:
    fixed_equipment_dates = fixed_equipment_dates or {}
    fixed_equipment = set(fixed_equipment_dates)
    passed_equipment = completed_equipment - failed_equipment
    module_records: list[dict[str, Any]] = []
    failed_items: list[dict[str, Any]] = []
    incomplete_items: list[dict[str, Any]] = []
    test_items: list[dict[str, Any]] = []
    matched_failed_keys: set[str] = set()

    for index, module_link in enumerate(module_links):
        module_key = compact_equipment_key(module_link.get("source_equipment_label"))
        equipment_info = lookup_equipment_info(module_link, info_by_key)
        module_link = link_with_effective_pdm(module_link, equipment_info)
        tracker_records = records_by_module.get(module_key, [])
        status, completed_records, failed_records, incomplete_records = (
            summarize_module_status(
                tracker_records,
                equipment_info,
                completed_equipment,
                failed_equipment,
                fixed_equipment,
            )
        )

        for record in tracker_records:
            retested_at = fixed_equipment_dates.get(record.equipment_key)
            if record_is_failed(
                record,
                failed_equipment,
                fixed_equipment,
                passed_equipment,
            ):
                item_status = STATUS_FAILED
            elif record_is_complete(record, completed_equipment):
                item_status = ITEM_STATUS_FIXED if retested_at else ITEM_STATUS_PASSED
            else:
                item_status = "Not Tested"

            test_items.append(
                item_record(
                    module_link=module_link,
                    equipment_info=equipment_info,
                    tracker_record=record,
                    item_status=item_status,
                    retested_at=retested_at,
                )
            )

        for record in failed_records:
            matched_failed_keys.add(record.equipment_key)
            failed_items.append(
                item_record(
                    module_link=module_link,
                    equipment_info=equipment_info,
                    tracker_record=record,
                    item_status=STATUS_FAILED,
                )
            )

        if status == STATUS_PARTIAL:
            for record in incomplete_records:
                incomplete_items.append(
                    item_record(
                        module_link=module_link,
                        equipment_info=equipment_info,
                        tracker_record=record,
                        item_status="Incomplete",
                    )
                )

        tracker_item_count = len(tracker_records)
        completed_count = len(completed_records)
        module_records.append(
            {
                "row_id": f"{module_link.get('pdm_name')}-{module_key}-{index}",
                "pdm_name": module_link.get("pdm_name"),
                "module_equipment": module_link.get("source_equipment_label"),
                "module_equipment_key": module_key,
                "matched_equipment_id": module_link.get("matched_equipment_id"),
                "match_status": module_link.get("match_status"),
                "source_equipment_column": module_link.get("source_equipment_column"),
                "eps_test_status": status,
                "tracker_item_count": tracker_item_count,
                "completed_test_item_count": completed_count,
                "incomplete_test_item_count": len(incomplete_records),
                "failed_test_item_count": len(failed_records),
                "field_test_completion_rate": (
                    completed_count / tracker_item_count if tracker_item_count else None
                ),
                "neta_complete": equipment_info.get("neta_complete"),
                "neta_test_report": equipment_info.get("neta_test_report"),
                "equipment_serial_number": equipment_info.get("serial_number"),
                "equipment_manufacturer": equipment_info.get("manufacturer"),
                "equipment_model": equipment_info.get("model"),
                "tracker_types": sorted(
                    {record.test_type for record in tracker_records if record.test_type}
                ),
            }
        )

    return (
        module_records,
        failed_items,
        incomplete_items,
        test_items,
        matched_failed_keys,
    )


def add_unmatched_failed_items(
    failed_items: list[dict[str, Any]],
    failed_equipment: set[str],
    tracker_equipment_keys: set[str],
    matched_failed_keys: set[str],
    tracker_by_substation: dict[str, list[TrackerRecord]],
    info_by_key: dict[str, dict[str, Any]],
) -> None:
    records_by_equipment: dict[str, list[TrackerRecord]] = defaultdict(list)
    for records in tracker_by_substation.values():
        for record in records:
            records_by_equipment[record.equipment_key].append(record)

    for equipment_key in sorted(
        (failed_equipment & tracker_equipment_keys) - matched_failed_keys
    ):
        for record in records_by_equipment[equipment_key]:
            equipment_info = info_by_key.get(
                compact_equipment_key(record.substation), {}
            )
            failed_items.append(
                {
                    "item_status": STATUS_FAILED,
                    "pdm_name": "Not in module list",
                    "module_equipment": record.substation,
                    "module_equipment_key": compact_equipment_key(record.substation),
                    "matched_equipment_id": None,
                    "equipment_name": record.equipment_name,
                    "equipment_key": record.equipment_key,
                    "tracker_row": record.row_number,
                    "tracker_type": record.test_type,
                    "tracker_equipment_type": record.equipment_type,
                    "follow_up_req": record.follow_up_req,
                    "comments": record.comments
                    or "No matching equipment cell in module list.",
                    "date_tested": record.date_tested,
                    "equipment_serial_number": equipment_info.get("serial_number"),
                    "equipment_manufacturer": equipment_info.get("manufacturer"),
                    "equipment_model": equipment_info.get("model"),
                    "neta_complete": equipment_info.get("neta_complete"),
                    "neta_test_report": equipment_info.get("neta_test_report"),
                }
            )


def pdm_execution_status(status_counts: Counter[str]) -> str:
    if status_counts[STATUS_FAILED] > 0:
        return STATUS_FAILED
    if status_counts[STATUS_PARTIAL] > 0:
        return STATUS_PARTIAL
    if status_counts[STATUS_WAITING_INFRALINK_NETA] > 0:
        return STATUS_WAITING_INFRALINK_NETA
    if status_counts[STATUS_COMPLETE] > 0 and (
        status_counts[STATUS_NOT_STARTED] == 0
        and status_counts[STATUS_NO_TRACKER_RECORDS] == 0
    ):
        return STATUS_COMPLETE
    if (
        status_counts[STATUS_COMPLETE] > 0
        or status_counts[STATUS_WAITING_INFRALINK_NETA] > 0
    ):
        return STATUS_PARTIAL
    if (
        status_counts[STATUS_NO_TRACKER_RECORDS] > 0
        and status_counts[STATUS_NOT_STARTED] == 0
    ):
        return STATUS_NO_TRACKER_RECORDS
    return STATUS_NOT_STARTED


def build_pdm_execution_records(
    module_records: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for record in module_records:
        grouped[clean(record.get("pdm_name")) or "Unknown PDM"].append(record)

    pdm_records: list[dict[str, Any]] = []
    for pdm_name, records in grouped.items():
        counts = Counter(record["eps_test_status"] for record in records)
        total = len(records)
        started_count = (
            total - counts[STATUS_NOT_STARTED] - counts[STATUS_NO_TRACKER_RECORDS]
        )
        tracker_items = sum(
            int(record.get("tracker_item_count") or 0) for record in records
        )
        completed_items = sum(
            int(record.get("completed_test_item_count") or 0) for record in records
        )
        failed_items = sum(
            int(record.get("failed_test_item_count") or 0) for record in records
        )

        pdm_records.append(
            {
                "pdm_name": pdm_name,
                "eps_execution_status": pdm_execution_status(counts),
                "module_equipment_count": total,
                "started_module_equipment_count": started_count,
                "complete_count": counts[STATUS_COMPLETE],
                "waiting_infralink_neta_count": counts[STATUS_WAITING_INFRALINK_NETA],
                "partial_count": counts[STATUS_PARTIAL],
                "failed_count": counts[STATUS_FAILED],
                "not_started_count": counts[STATUS_NOT_STARTED],
                "no_tracker_record_count": counts[STATUS_NO_TRACKER_RECORDS],
                "tracker_item_count": tracker_items,
                "completed_test_item_count": completed_items,
                "failed_test_item_count": failed_items,
                "field_test_completion_rate": (
                    completed_items / tracker_items if tracker_items else None
                ),
            }
        )

    return sorted(
        pdm_records,
        key=lambda record: (
            STATUS_ORDER.get(str(record["eps_execution_status"]), 99),
            -int(record.get("failed_count") or 0),
            -int(record.get("partial_count") or 0),
            -int(record.get("waiting_infralink_neta_count") or 0),
            str(record.get("pdm_name") or ""),
        ),
    )


def source_files_metadata() -> dict[str, dict[str, str] | None]:
    tracker_source = find_latest_tracker_source_file()
    tracker_input = get_tracker_input_path()
    sources = {
        "eps_tracker": tracker_input,
        "eps_tracker_workbook": EPS_TRACKER_PATH,
        "eps_tracker_source": tracker_source,
        "daily_tested_equipment": DAILY_TESTED_EQUIPMENT_PATH,
        "module_equipment_links": MODULE_LINKS_PATH,
        "equipment": EQUIPMENT_PATH,
    }
    return {
        name: file_metadata(path) if path and Path(path).exists() else None
        for name, path in sources.items()
    }


def find_latest_tracker_source_file() -> Path | None:
    if not EPS_TRACKER_SOURCE_DIR.exists():
        return None
    candidates = [
        path
        for path in EPS_TRACKER_SOURCE_DIR.glob("*.xlsx")
        if path.is_file() and not path.name.startswith("~$")
    ]
    if not candidates:
        return None
    return max(candidates, key=lambda path: (path.stat().st_mtime, path.name))


def get_tracker_input_path() -> Path:
    return find_latest_tracker_source_file() or EPS_TRACKER_PATH


def snapshot_date_today() -> date:
    return date.today()


def snapshot_path_for(snapshot_date: date) -> Path:
    return SNAPSHOT_DIR / f"{snapshot_date.isoformat()}.json"


def parse_snapshot_date(path: Path) -> date | None:
    try:
        return date.fromisoformat(path.stem)
    except ValueError:
        return None


def load_snapshot(path: Path | None) -> dict[str, Any] | None:
    if path is None or not path.exists():
        return None
    import json

    return json.loads(path.read_text(encoding="utf-8"))


def list_snapshot_paths(current_date: date) -> list[tuple[date, Path]]:
    if not SNAPSHOT_DIR.exists():
        return []

    snapshots: list[tuple[date, Path]] = []
    for path in SNAPSHOT_DIR.glob("*.json"):
        snapshot_date = parse_snapshot_date(path)
        if snapshot_date is None or snapshot_date >= current_date:
            continue
        snapshots.append((snapshot_date, path))
    return sorted(snapshots)


def find_latest_snapshot_before(current_date: date) -> Path | None:
    snapshots = list_snapshot_paths(current_date)
    return snapshots[-1][1] if snapshots else None


def choose_baseline_snapshot(
    snapshots: list[tuple[date, Path]],
    current_date: date,
    lookback_days: int,
    tolerance_days: int = 2,
) -> Path | None:
    target_date = current_date - timedelta(days=lookback_days)

    sorted_snapshots = sorted(snapshots)
    older_or_target = [
        (snapshot_date, path)
        for snapshot_date, path in sorted_snapshots
        if snapshot_date <= target_date
    ]
    if older_or_target:
        return older_or_target[-1][1]

    min_age_days = max(1, lookback_days - tolerance_days)
    max_age_days = lookback_days + tolerance_days
    fallback_candidates = [
        (snapshot_date, path)
        for snapshot_date, path in sorted_snapshots
        if min_age_days <= (current_date - snapshot_date).days <= max_age_days
    ]
    if not fallback_candidates:
        return None

    return min(
        fallback_candidates,
        key=lambda item: abs((item[0] - target_date).days),
    )[1]


def find_baseline_snapshot(current_date: date, lookback_days: int) -> Path | None:
    return choose_baseline_snapshot(
        list_snapshot_paths(current_date),
        current_date,
        lookback_days,
    )


def make_snapshot(
    snapshot_date: date,
    source_date: date,
    passed_equipment: set[str],
    failed_equipment: set[str],
    fixed_equipment: set[str],
    module_records: list[dict[str, Any]],
    pdm_records: list[dict[str, Any]],
) -> dict[str, Any]:
    return {
        "snapshot_date": snapshot_date.isoformat(),
        "source_date_label": source_date.isoformat(),
        "tested_equipment": sorted(passed_equipment),
        "failed_equipment": sorted(failed_equipment),
        "fixed_equipment": sorted(fixed_equipment),
        "module_status_by_key": {
            str(record["module_equipment_key"]): record["eps_test_status"]
            for record in module_records
            if record.get("module_equipment_key")
        },
        "pdm_status_by_name": {
            str(record["pdm_name"]): {
                "eps_execution_status": record.get("eps_execution_status"),
                "complete_count": record.get("complete_count", 0),
                "waiting_infralink_neta_count": record.get(
                    "waiting_infralink_neta_count", 0
                ),
                "partial_count": record.get("partial_count", 0),
                "failed_count": record.get("failed_count", 0),
                "not_started_count": record.get("not_started_count", 0),
                "no_tracker_record_count": record.get("no_tracker_record_count", 0),
            }
            for record in pdm_records
            if record.get("pdm_name")
        },
    }


def build_snapshot_for_daily_source_date(
    *,
    history: DailyTestedEquipmentHistory,
    source_date: date,
    snapshot_date: date,
    tracker_equipment_keys: set[str],
    module_links: list[dict[str, Any]],
    info_by_key: dict[str, dict[str, Any]],
    records_by_module: dict[str, list[TrackerRecord]],
) -> dict[str, Any]:
    passed_input_equipment, failed_input_equipment = cumulative_daily_equipment(
        history,
        source_date,
    )
    retested_input_dates = current_retested_equipment_dates(
        history,
        source_date,
        passed_input_equipment,
    )
    equipment_sets = build_execution_equipment_sets(
        passed_input_equipment,
        failed_input_equipment,
        tracker_equipment_keys,
    )
    fixed_equipment_dates = tracker_retested_equipment_dates(
        retested_input_dates,
        tracker_equipment_keys,
    )
    fixed_snapshot_equipment = build_execution_equipment_sets(
        set(retested_input_dates),
        set(),
        tracker_equipment_keys,
    ).snapshot_passed_equipment
    module_records, *_rest = build_module_execution_records(
        module_links,
        info_by_key,
        records_by_module,
        equipment_sets.completed_equipment,
        equipment_sets.failed_equipment,
        fixed_equipment_dates,
    )
    pdm_records = build_pdm_execution_records(module_records)
    return make_snapshot(
        snapshot_date,
        source_date,
        equipment_sets.snapshot_passed_equipment,
        equipment_sets.snapshot_failed_equipment,
        fixed_snapshot_equipment,
        module_records,
        pdm_records,
    )


def write_daily_history_snapshots(
    *,
    history: DailyTestedEquipmentHistory,
    tracker_equipment_keys: set[str],
    module_links: list[dict[str, Any]],
    info_by_key: dict[str, dict[str, Any]],
    records_by_module: dict[str, list[TrackerRecord]],
) -> None:
    for source_date in history.dates:
        snapshot = build_snapshot_for_daily_source_date(
            history=history,
            source_date=source_date,
            snapshot_date=source_date + timedelta(days=1),
            tracker_equipment_keys=tracker_equipment_keys,
            module_links=module_links,
            info_by_key=info_by_key,
            records_by_module=records_by_module,
        )
        write_json(snapshot_path_for(source_date + timedelta(days=1)), snapshot)


def compare_snapshots(
    current: dict[str, Any],
    baseline: dict[str, Any] | None,
    lookback_days: int,
) -> dict[str, Any]:
    current_label = current.get("source_date_label") or current.get("snapshot_date")
    if baseline is None:
        return {
            "available": False,
            "current_date": current_label,
            "baseline_date": None,
            "target_days": lookback_days,
            "new_tested_count": 0,
            "new_failed_count": 0,
            "repaired_count": 0,
            "new_complete_module_count": 0,
            "new_waiting_infralink_neta_count": 0,
            "new_failed_module_count": 0,
            "new_tested_equipment": [],
            "new_failed_equipment": [],
            "repaired_equipment": [],
        }

    baseline_label = baseline.get("source_date_label") or baseline.get("snapshot_date")
    current_tested = set(current.get("tested_equipment") or [])
    baseline_tested = set(baseline.get("tested_equipment") or [])
    current_failed = set(current.get("failed_equipment") or [])
    baseline_failed = set(baseline.get("failed_equipment") or [])
    current_fixed = set(current.get("fixed_equipment") or [])
    baseline_fixed = set(baseline.get("fixed_equipment") or [])
    current_module_status = current.get("module_status_by_key") or {}
    baseline_module_status = baseline.get("module_status_by_key") or {}

    new_tested = sorted(current_tested - baseline_tested)
    new_failed = sorted(current_failed - baseline_failed)
    repaired = sorted(
        ((baseline_failed - current_failed) & current_tested)
        | (current_fixed - baseline_fixed)
    )
    new_complete_modules = sorted(
        module_key
        for module_key, status in current_module_status.items()
        if status == STATUS_COMPLETE
        and baseline_module_status.get(module_key) != STATUS_COMPLETE
    )
    new_waiting_modules = sorted(
        module_key
        for module_key, status in current_module_status.items()
        if status == STATUS_WAITING_INFRALINK_NETA
        and baseline_module_status.get(module_key) != STATUS_WAITING_INFRALINK_NETA
    )
    new_failed_modules = sorted(
        module_key
        for module_key, status in current_module_status.items()
        if status == STATUS_FAILED
        and baseline_module_status.get(module_key) != STATUS_FAILED
    )

    return {
        "available": True,
        "current_date": current_label,
        "baseline_date": baseline_label,
        "target_days": lookback_days,
        "new_tested_count": len(new_tested),
        "new_failed_count": len(new_failed),
        "repaired_count": len(repaired),
        "current_tested_count": len(current_tested),
        "baseline_tested_count": len(baseline_tested),
        "current_failed_count": len(current_failed),
        "baseline_failed_count": len(baseline_failed),
        "new_complete_module_count": len(new_complete_modules),
        "new_waiting_infralink_neta_count": len(new_waiting_modules),
        "new_failed_module_count": len(new_failed_modules),
        "new_tested_equipment": new_tested,
        "new_failed_equipment": new_failed,
        "repaired_equipment": repaired,
        "new_complete_modules": new_complete_modules,
        "new_waiting_infralink_neta_modules": new_waiting_modules,
        "new_failed_modules": new_failed_modules,
    }


def is_one_day_snapshot_diff(comparison: dict[str, Any], current_date: date) -> bool:
    if not comparison.get("available"):
        return False

    try:
        baseline_date = date.fromisoformat(str(comparison.get("baseline_date")))
    except ValueError:
        return False

    return baseline_date == current_date - timedelta(days=1)


def build_summary(
    *,
    current_snapshot: dict[str, Any],
    seven_day_baseline: dict[str, Any] | None,
    previous_seven_day_baseline: dict[str, Any] | None,
    yesterday_summary: dict[str, Any],
    daily_history_summary: dict[str, Any],
    passed_equipment: set[str],
    failed_equipment: set[str],
    module_records: list[dict[str, Any]],
    pdm_records: list[dict[str, Any]],
    failed_items: list[dict[str, Any]],
    incomplete_items: list[dict[str, Any]],
    test_items: list[dict[str, Any]],
    not_found_items: list[dict[str, Any]],
    alias_matches: dict[str, str],
) -> dict[str, Any]:
    status_counts = Counter(record["eps_test_status"] for record in module_records)
    tracker_item_count = sum(
        1
        for item in test_items
        if "Not In Tracker" not in str(item.get("item_status") or "")
    )
    completed_item_count = sum(
        1
        for item in test_items
        if item.get("item_status")
        in {STATUS_FAILED, ITEM_STATUS_PASSED, ITEM_STATUS_FIXED}
    )
    yesterday = {
        **yesterday_summary,
        "snapshot_diff_available": False,
        "current_date": current_snapshot.get("source_date_label"),
        "baseline_date": None,
    }

    seven_day = compare_snapshots(current_snapshot, seven_day_baseline, LOOKBACK_DAYS)
    previous_seven_day = (
        compare_snapshots(
            seven_day_baseline,
            previous_seven_day_baseline,
            LOOKBACK_DAYS,
        )
        if seven_day_baseline
        else {"available": False}
    )

    return {
        "generated_at": datetime.now().astimezone().isoformat(),
        "selected_input_files": source_files_metadata(),
        "snapshot_date": current_snapshot.get("snapshot_date"),
        "source_date_label": current_snapshot.get("source_date_label"),
        "total_pdm_count": len(pdm_records),
        "total_module_equipment_count": len(module_records),
        "total_tracker_test_item_count": tracker_item_count,
        "completed_tracker_test_item_count": completed_item_count,
        "field_test_completion_rate": (
            completed_item_count / tracker_item_count if tracker_item_count else None
        ),
        "total_tested_equipment_count": len(passed_equipment),
        "total_failed_equipment_count": len(failed_equipment),
        "failed_test_item_count": sum(
            1
            for item in test_items
            if item.get("item_status") in {STATUS_FAILED, "Failed - Not In Tracker"}
        ),
        "incomplete_test_item_count": sum(
            1 for item in test_items if item.get("item_status") == "Not Tested"
        ),
        "test_item_count": len(test_items),
        "passed_test_item_count": sum(
            1
            for item in test_items
            if item.get("item_status")
            in {
                ITEM_STATUS_PASSED,
                ITEM_STATUS_PASSED_NOT_IN_TRACKER,
                ITEM_STATUS_FIXED,
                ITEM_STATUS_FIXED_NOT_IN_TRACKER,
            }
        ),
        "fixed_test_item_count": sum(
            1
            for item in test_items
            if item.get("item_status")
            in {ITEM_STATUS_FIXED, ITEM_STATUS_FIXED_NOT_IN_TRACKER}
        ),
        "tested_test_item_count": sum(
            1
            for item in test_items
            if item.get("item_status")
            in {
                ITEM_STATUS_PASSED,
                ITEM_STATUS_PASSED_NOT_IN_TRACKER,
                ITEM_STATUS_FIXED,
                ITEM_STATUS_FIXED_NOT_IN_TRACKER,
            }
        ),
        "not_tested_test_item_count": sum(
            1 for item in test_items if item.get("item_status") == "Not Tested"
        ),
        "not_found_test_item_count": len(not_found_items),
        "pdu_feeder_breaker_alias_match_count": len(alias_matches),
        "status_counts": dict(status_counts),
        "complete_count": status_counts[STATUS_COMPLETE],
        "waiting_infralink_neta_count": status_counts[STATUS_WAITING_INFRALINK_NETA],
        "partial_count": status_counts[STATUS_PARTIAL],
        "failed_count": status_counts[STATUS_FAILED],
        "not_started_count": status_counts[STATUS_NOT_STARTED],
        "no_tracker_record_count": status_counts[STATUS_NO_TRACKER_RECORDS],
        "yesterday": yesterday,
        "seven_day": seven_day,
        "previous_seven_day": previous_seven_day,
        "daily_history": daily_history_summary,
        "top_pdms_by_failed": sorted(
            [
                record
                for record in pdm_records
                if int(record.get("failed_count") or 0) > 0
            ],
            key=lambda record: (
                -int(record.get("failed_count") or 0),
                str(record.get("pdm_name")),
            ),
        )[:10],
        "top_pdms_by_partial": sorted(
            [
                record
                for record in pdm_records
                if int(record.get("partial_count") or 0) > 0
            ],
            key=lambda record: (
                -int(record.get("partial_count") or 0),
                str(record.get("pdm_name")),
            ),
        )[:10],
        "top_pdms_waiting_infralink_neta": sorted(
            [
                record
                for record in pdm_records
                if int(record.get("waiting_infralink_neta_count") or 0) > 0
            ],
            key=lambda record: (
                -int(record.get("waiting_infralink_neta_count") or 0),
                str(record.get("pdm_name")),
            ),
        )[:10],
    }


def records_envelope(records: list[dict[str, Any]]) -> dict[str, Any]:
    return {
        "generated_at": datetime.now().astimezone().isoformat(),
        "selected_input_files": source_files_metadata(),
        "records": records,
    }


def build_eps_test_execution() -> dict[str, Any]:
    tracker_input_path = get_tracker_input_path()
    missing_inputs = [
        path
        for path in [
            tracker_input_path,
            DAILY_TESTED_EQUIPMENT_PATH,
            MODULE_LINKS_PATH,
            EQUIPMENT_PATH,
        ]
        if not path.exists()
    ]
    if missing_inputs:
        raise FileNotFoundError(
            "Missing EPS execution input files: "
            + ", ".join(str(path) for path in missing_inputs)
        )

    current_date = snapshot_date_today()
    daily_history = parse_daily_tested_equipment(
        DAILY_TESTED_EQUIPMENT_PATH,
        reference_date=current_date,
    )
    current_source_date = daily_history.latest_date
    if current_source_date is None:
        raise ValueError(
            f"No dated equipment records found in {DAILY_TESTED_EQUIPMENT_PATH}"
        )

    passed_input_equipment, failed_input_equipment = cumulative_daily_equipment(
        daily_history,
        current_source_date,
    )
    retested_input_dates = current_retested_equipment_dates(
        daily_history,
        current_source_date,
        passed_input_equipment,
    )
    tracker_by_substation, tracker_equipment_keys = load_tracker_records(
        tracker_input_path
    )
    equipment_sets = build_execution_equipment_sets(
        passed_input_equipment,
        failed_input_equipment,
        tracker_equipment_keys,
    )
    fixed_equipment_dates = tracker_retested_equipment_dates(
        retested_input_dates,
        tracker_equipment_keys,
    )
    fixed_snapshot_equipment = build_execution_equipment_sets(
        set(retested_input_dates),
        set(),
        tracker_equipment_keys,
    ).snapshot_passed_equipment
    not_found_items = build_not_found_test_items(
        passed_input_equipment,
        failed_input_equipment,
        tracker_equipment_keys,
        retested_input_dates,
    )

    module_links = load_module_links()
    module_keys = {
        compact_equipment_key(link.get("source_equipment_label"))
        for link in module_links
    }
    module_keys = {module_key for module_key in module_keys if module_key}
    records_by_module = tracker_records_by_module(tracker_by_substation, module_keys)
    info_by_key = load_equipment_info()
    not_found_records = build_not_found_records(
        not_found_items,
        module_links,
        module_keys,
        info_by_key,
    )
    (
        module_records,
        failed_items,
        incomplete_items,
        _module_test_items,
        matched_failed_keys,
    ) = build_module_execution_records(
        module_links,
        info_by_key,
        records_by_module,
        equipment_sets.completed_equipment,
        equipment_sets.failed_equipment,
        fixed_equipment_dates,
    )
    add_unmatched_failed_items(
        failed_items,
        equipment_sets.failed_equipment,
        tracker_equipment_keys,
        matched_failed_keys,
        tracker_by_substation,
        info_by_key,
    )
    test_items = build_tracker_test_items(
        tracker_by_substation,
        module_links,
        module_keys,
        info_by_key,
        equipment_sets.completed_equipment,
        equipment_sets.failed_equipment,
        fixed_equipment_dates,
    )
    test_items.extend(not_found_records)
    pdm_records = build_pdm_execution_records(module_records)

    current_snapshot = make_snapshot(
        current_date,
        current_source_date,
        equipment_sets.snapshot_passed_equipment,
        equipment_sets.snapshot_failed_equipment,
        fixed_snapshot_equipment,
        module_records,
        pdm_records,
    )
    seven_day_source_date = choose_baseline_daily_date(
        daily_history.dates,
        current_source_date,
        LOOKBACK_DAYS,
    )
    seven_day_baseline = (
        build_snapshot_for_daily_source_date(
            history=daily_history,
            source_date=seven_day_source_date,
            snapshot_date=seven_day_source_date + timedelta(days=1),
            tracker_equipment_keys=tracker_equipment_keys,
            module_links=module_links,
            info_by_key=info_by_key,
            records_by_module=records_by_module,
        )
        if seven_day_source_date
        else None
    )
    previous_seven_day_source_date = (
        choose_baseline_daily_date(
            daily_history.dates,
            seven_day_source_date,
            LOOKBACK_DAYS,
        )
        if seven_day_source_date
        else None
    )
    previous_seven_day_baseline = (
        build_snapshot_for_daily_source_date(
            history=daily_history,
            source_date=previous_seven_day_source_date,
            snapshot_date=previous_seven_day_source_date + timedelta(days=1),
            tracker_equipment_keys=tracker_equipment_keys,
            module_links=module_links,
            info_by_key=info_by_key,
            records_by_module=records_by_module,
        )
        if previous_seven_day_source_date
        else None
    )
    yesterday_summary = daily_activity_summary(
        daily_history,
        current_source_date,
        tracker_equipment_keys,
    )
    daily_history_summary = build_daily_history_summary(
        daily_history,
        tracker_equipment_keys,
    )

    summary = build_summary(
        current_snapshot=current_snapshot,
        seven_day_baseline=seven_day_baseline,
        previous_seven_day_baseline=previous_seven_day_baseline,
        yesterday_summary=yesterday_summary,
        daily_history_summary=daily_history_summary,
        passed_equipment=equipment_sets.snapshot_passed_equipment,
        failed_equipment=equipment_sets.snapshot_failed_equipment,
        module_records=module_records,
        pdm_records=pdm_records,
        failed_items=failed_items,
        incomplete_items=incomplete_items,
        test_items=test_items,
        not_found_items=not_found_records,
        alias_matches=equipment_sets.alias_matches,
    )

    write_json(SUMMARY_OUTPUT_PATH, summary)
    write_json(PDM_OUTPUT_PATH, records_envelope(pdm_records))
    write_json(MODULE_OUTPUT_PATH, records_envelope(module_records))
    write_json(TEST_ITEMS_OUTPUT_PATH, records_envelope(test_items))
    write_json(FAILED_OUTPUT_PATH, records_envelope(failed_items))
    write_json(INCOMPLETE_OUTPUT_PATH, records_envelope(incomplete_items))
    write_json(NOT_FOUND_OUTPUT_PATH, records_envelope(not_found_records))
    write_daily_history_snapshots(
        history=daily_history,
        tracker_equipment_keys=tracker_equipment_keys,
        module_links=module_links,
        info_by_key=info_by_key,
        records_by_module=records_by_module,
    )
    write_json(snapshot_path_for(current_date), current_snapshot)

    return summary


def run_build(input_files: dict[str, str] | None = None) -> dict[str, Any]:
    _ = input_files
    return build_eps_test_execution()


def main() -> None:
    summary = build_eps_test_execution()
    print(f"Wrote EPS test execution summary to {SUMMARY_OUTPUT_PATH}")
    print(f"EPS module equipment records: {summary['total_module_equipment_count']}")
    print(f"EPS failed module equipment: {summary['failed_count']}")
    print(
        "Complete, waiting Infralink NETA completion: "
        f"{summary['waiting_infralink_neta_count']}"
    )


if __name__ == "__main__":
    main()
